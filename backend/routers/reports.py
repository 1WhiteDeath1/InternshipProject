"""Reports and analytics router - the Manager's cross-module reporting
surface (permission: reports:view, held only by Manager/Deputy Manager).

Each JSON endpoint's computation lives in a plain `_xxx_data(db, ...)`
helper so /export can reuse the exact same numbers it renders on screen,
rather than recomputing them a second, divergent way."""
import io
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from backend.database import get_db
from backend.models import (
    Invoice, InvoiceStatus, Booking, Room, InventoryItem, StockBatch, WasteLog,
    Vendor, AuditLog, Alert, AlertStatus, KitchenOrder,
    IncidentReport, MealAttendance, AttendanceStatus, MessBill, MessBillStatus,
    Member, MemberStatus, InvoiceEditRequest, MenuItemEditRequest, EditRequestStatus,
)
from backend.auth import PermissionChecker

router = APIRouter()


# --- Shared helpers -------------------------------------------------------

def _invoice_revenue(db: Session, start_dt: datetime, end_dt: datetime, bill_type: str | None = None):
    q = db.query(func.sum(Invoice.total_amount)).filter(
        Invoice.created_at >= start_dt, Invoice.created_at < end_dt,
        Invoice.status.in_(["issued", "paid"]),
    )
    if bill_type:
        q = q.filter(Invoice.bill_type == bill_type)
    return float(q.scalar() or 0)


def _period_cost(db: Session, start_dt: datetime, end_dt: datetime):
    """Real, trackable operating cost for the period: self-purchase stock
    spend (raw materials, logged via Daily Stock Intake / Smart Intake)
    plus logged waste. No payroll/utilities module exists to draw from."""
    procurement = float(db.query(func.sum(StockBatch.quantity * StockBatch.unit_cost)).filter(
        StockBatch.created_at >= start_dt, StockBatch.created_at < end_dt,
    ).scalar() or 0)
    waste = float(db.query(func.sum(WasteLog.cost)).filter(
        WasteLog.created_at >= start_dt, WasteLog.created_at < end_dt,
    ).scalar() or 0)
    return procurement, waste


def _pct(curr: float, prev: float) -> float:
    if prev > 0:
        return round((curr - prev) / prev * 100, 1)
    return 100.0 if curr > 0 else 0.0


# --- KPI snapshot ("This Month at a Glance") ------------------------------

def _dashboard_kpis(db: Session) -> dict:
    today = date.today()
    month_start = today.replace(day=1)

    invoices = db.query(Invoice).filter(Invoice.created_at >= datetime.combine(today, datetime.min.time())).all()
    today_revenue = sum(float(inv.total_amount) for inv in invoices if inv.status.value in ["issued", "paid"])

    total_rooms = db.query(Room).filter(Room.is_active == True).count()
    occupied = db.query(Room).filter(Room.status == "occupied").count()
    occupancy_rate = round(occupied / total_rooms * 100, 1) if total_rooms else 0

    batches = db.query(StockBatch).filter(StockBatch.is_active == True).all()
    stock_value = sum(float(b.unit_cost) * b.quantity for b in batches if b.unit_cost)

    waste_items = db.query(WasteLog).filter(WasteLog.created_at >= datetime.combine(month_start, datetime.min.time())).all()
    waste_cost = sum(float(w.cost) for w in waste_items if w.cost)

    open_alerts = db.query(Alert).filter(Alert.status.in_(["new", "acknowledged"])).count()

    live_invoices = db.query(Invoice).filter(Invoice.status.in_(["draft", "issued"])).all()
    outstanding_balance = sum(max(float(inv.total_amount) - float(inv.amount_paid or 0), 0.0) for inv in live_invoices)
    unsettled_invoice_count = sum(1 for inv in live_invoices if float(inv.total_amount) - float(inv.amount_paid or 0) > 0.01)

    today_guests = db.query(Booking).filter(
        Booking.check_in <= today, Booking.check_out >= today, Booking.status.in_(["checked_in"]),
    ).count()

    from sqlalchemy import text
    low_stock_result = db.execute(text("""
        SELECT COUNT(*) FROM (
            SELECT i.id FROM inventory_items i
            LEFT JOIN stock_batches b ON b.item_id = i.id AND b.is_active = 1
            WHERE i.is_active = 1 AND i.reorder_level > 0
            GROUP BY i.id
            HAVING COALESCE(SUM(b.quantity), 0) <= i.reorder_level
        )
    """)).scalar()

    open_incidents = db.query(IncidentReport).filter(IncidentReport.status.in_(["open", "investigating"])).count()

    today_attendance = db.query(MealAttendance).filter(MealAttendance.date == today).all()
    attendance_present = sum(1 for a in today_attendance if a.status == AttendanceStatus.ATTENDED)
    attendance_absent = sum(1 for a in today_attendance if a.status == AttendanceStatus.NO_SHOW)

    active_vendor_count = db.query(Vendor).filter(Vendor.is_active == True).count()

    month_bills = db.query(MessBill).filter(MessBill.month == today.month, MessBill.year == today.year).all()
    mess_revenue_month = sum(float(b.total_amount) for b in month_bills)
    unpaid_mess_bills = sum(1 for b in month_bills if b.status != MessBillStatus.PAID)

    active_member_count = db.query(Member).filter(Member.status == MemberStatus.ACTIVE).count()

    invoices_finalized_today = sum(1 for inv in invoices if inv.status.value in ["issued", "paid"])
    month_invoices = db.query(Invoice).filter(Invoice.created_at >= datetime.combine(month_start, datetime.min.time())).all()
    discounts_month = sum(float(inv.discount) for inv in month_invoices if inv.discount)

    return {
        "today_revenue": today_revenue,
        "occupancy_rate": occupancy_rate,
        "total_stock_value": round(stock_value, 2),
        "waste_cost_month": round(waste_cost, 2),
        "open_alerts": open_alerts,
        "total_guests_today": today_guests,
        "low_stock_count": low_stock_result or 0,
        "open_incidents": open_incidents,
        "attendance_present_today": attendance_present,
        "attendance_absent_today": attendance_absent,
        "active_vendor_count": active_vendor_count,
        "mess_revenue_month": round(mess_revenue_month, 2),
        "unpaid_mess_bills": unpaid_mess_bills,
        "active_member_count": active_member_count,
        "invoices_finalized_today": invoices_finalized_today,
        "discounts_month": round(discounts_month, 2),
        "outstanding_balance": round(outstanding_balance, 2),
        "unsettled_invoice_count": unsettled_invoice_count,
    }


@router.get("/dashboard")
async def supervisor_dashboard(db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    return _dashboard_kpis(db)


# --- Revenue --------------------------------------------------------------

@router.get("/revenue-trend")
async def revenue_trend(days: int = Query(30, ge=7, le=365), db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    today = date.today()
    labels = []
    values = []
    for i in range(days - 1, -1, -1):
        d = today - timedelta(days=i)
        labels.append(d.strftime("%Y-%m-%d"))
        day_invoices = db.query(Invoice).filter(
            Invoice.created_at >= datetime.combine(d, datetime.min.time()),
            Invoice.created_at < datetime.combine(d + timedelta(days=1), datetime.min.time()),
            Invoice.status.in_(["issued", "paid"]),
        ).all()
        values.append(sum(float(inv.total_amount) for inv in day_invoices))
    return {"labels": labels, "values": values}


def _occupancy_trend_data(db: Session, days: int) -> dict:
    today = date.today()
    total_rooms = db.query(Room).filter(Room.is_active == True).count()

    def series(end_date: date) -> list[float]:
        vals = []
        for i in range(days - 1, -1, -1):
            d = end_date - timedelta(days=i)
            occupied = db.query(Booking).filter(
                Booking.check_in <= d, Booking.check_out >= d, Booking.status.in_(["checked_in"]),
            ).count()
            vals.append(round(occupied / total_rooms * 100, 1) if total_rooms else 0)
        return vals

    values = series(today)
    labels = [(today - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(days - 1, -1, -1)]
    prev_values = series(today - timedelta(days=days))

    avg_current = round(sum(values) / len(values), 1) if values else 0
    avg_previous = round(sum(prev_values) / len(prev_values), 1) if prev_values else 0

    return {
        "labels": labels, "values": values,
        "avg_current": avg_current, "avg_previous": avg_previous,
        "pct_change": _pct(avg_current, avg_previous),
    }


@router.get("/occupancy-trend")
async def occupancy_trend(days: int = Query(30, ge=7, le=365), db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    return _occupancy_trend_data(db, days)


@router.get("/revenue-summary")
async def revenue_summary(db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    """Compact Manager-dashboard widget: a short sparkline plus this-week vs
    last-week compared like-for-like (same number of elapsed days), so a
    Tuesday reading isn't compared against a full Sun-Sat prior week."""
    today = date.today()
    sparkline = []
    for i in range(13, -1, -1):
        d = today - timedelta(days=i)
        start = datetime.combine(d, datetime.min.time())
        sparkline.append(round(_invoice_revenue(db, start, start + timedelta(days=1)), 2))

    days_elapsed = today.weekday() + 1
    week_start = today - timedelta(days=today.weekday())
    last_week_start = week_start - timedelta(days=7)

    current_week_total = _invoice_revenue(db, datetime.combine(week_start, datetime.min.time()),
                                           datetime.combine(today + timedelta(days=1), datetime.min.time()))
    last_week_comparable_total = _invoice_revenue(
        db, datetime.combine(last_week_start, datetime.min.time()),
        datetime.combine(last_week_start + timedelta(days=days_elapsed), datetime.min.time()))

    return {
        "sparkline": sparkline,
        "current_week_total": round(current_week_total, 2),
        "last_week_comparable_total": round(last_week_comparable_total, 2),
        "pct_change": _pct(current_week_total, last_week_comparable_total),
    }


def _revenue_detail_data(db: Session, days: int) -> dict:
    today = date.today()
    period_end = today + timedelta(days=1)
    period_start = today - timedelta(days=days - 1)

    weekly = days > 90
    labels, revenue_values, cost_values, profit_values = [], [], [], []
    bucket_start = period_start
    while bucket_start <= today:
        bucket_len = 7 if weekly else 1
        bucket_end = min(bucket_start + timedelta(days=bucket_len), period_end)
        start_dt = datetime.combine(bucket_start, datetime.min.time())
        end_dt = datetime.combine(bucket_end, datetime.min.time())
        rev = _invoice_revenue(db, start_dt, end_dt)
        procurement, waste = _period_cost(db, start_dt, end_dt)
        cost = procurement + waste
        labels.append(bucket_start.strftime("%Y-%m-%d"))
        revenue_values.append(round(rev, 2))
        cost_values.append(round(cost, 2))
        profit_values.append(round(rev - cost, 2))
        bucket_start = bucket_end

    full_start_dt = datetime.combine(period_start, datetime.min.time())
    full_end_dt = datetime.combine(period_end, datetime.min.time())
    room_revenue_total = _invoice_revenue(db, full_start_dt, full_end_dt, bill_type="room")
    mess_revenue_total = _invoice_revenue(db, full_start_dt, full_end_dt, bill_type="mess")
    revenue_total = sum(revenue_values)
    procurement_total, waste_total = _period_cost(db, full_start_dt, full_end_dt)
    cost_total = procurement_total + waste_total
    profit_total = revenue_total - cost_total
    other_revenue_total = round(revenue_total - room_revenue_total - mess_revenue_total, 2)

    prev_start = period_start - timedelta(days=days)
    prev_end = period_start
    prev_start_dt = datetime.combine(prev_start, datetime.min.time())
    prev_end_dt = datetime.combine(prev_end, datetime.min.time())
    prev_revenue = _invoice_revenue(db, prev_start_dt, prev_end_dt)
    prev_procurement, prev_waste = _period_cost(db, prev_start_dt, prev_end_dt)
    prev_cost = prev_procurement + prev_waste
    prev_profit = prev_revenue - prev_cost

    return {
        "labels": labels, "revenue": revenue_values, "cost": cost_values, "profit": profit_values,
        "room_revenue_total": round(room_revenue_total, 2),
        "mess_revenue_total": round(mess_revenue_total, 2),
        "other_revenue_total": other_revenue_total,
        "revenue_total": round(revenue_total, 2),
        "cost_total": round(cost_total, 2),
        "profit_total": round(profit_total, 2),
        "previous_revenue_total": round(prev_revenue, 2),
        "previous_cost_total": round(prev_cost, 2),
        "previous_profit_total": round(prev_profit, 2),
        "revenue_pct_change": _pct(revenue_total, prev_revenue),
        "cost_pct_change": _pct(cost_total, prev_cost),
        "profit_pct_change": _pct(profit_total, prev_profit),
    }


@router.get("/revenue-detail")
async def revenue_detail(days: int = Query(30, ge=7, le=365), db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    """Manager-dashboard revenue popup: a revenue/cost/profit series for the
    requested range (1w/1m/3m/1y, daily buckets except 1y which buckets
    weekly to keep the series readable), the room/mess split, and a
    previous-equal-length-period comparison."""
    return _revenue_detail_data(db, days)


@router.get("/occupancy-detail")
async def occupancy_detail(db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    """Manager-dashboard occupancy popup: today's occupied/non-occupied split
    plus a breakdown of each side, and last week's occupancy rate for the
    inner trend ring."""
    today = date.today()
    total_rooms = db.query(Room).filter(Room.is_active == True).count()

    non_occupied_by_status = {
        status: db.query(Room).filter(Room.is_active == True, Room.status == status).count()
        for status in ["vacant", "reserved", "maintenance"]
    }
    occupied_count = db.query(Room).filter(Room.is_active == True, Room.status == "occupied").count()
    not_occupied_count = sum(non_occupied_by_status.values())

    duty_rows = db.query(Booking.nature_of_duty, func.count(Booking.id)).filter(
        Booking.status == "checked_in",
    ).group_by(Booking.nature_of_duty).all()
    occupied_by_duty = {duty or "visit": count for duty, count in duty_rows}

    last_week = today - timedelta(days=7)
    last_week_occupied = db.query(Booking).filter(
        Booking.check_in <= last_week, Booking.check_out >= last_week,
        Booking.status.in_(["checked_in"]),
    ).count()
    last_week_rate = round(last_week_occupied / total_rooms * 100, 1) if total_rooms else 0

    return {
        "total_rooms": total_rooms,
        "occupied_count": occupied_count,
        "not_occupied_count": not_occupied_count,
        "occupancy_rate": round(occupied_count / total_rooms * 100, 1) if total_rooms else 0,
        "non_occupied_by_status": non_occupied_by_status,
        "occupied_by_duty": occupied_by_duty,
        "last_week_occupancy_rate": last_week_rate,
    }


@router.get("/cost-revenue-flow")
async def cost_revenue_flow(days: int = Query(30, ge=1, le=365), db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    """Manager-dashboard Sankey source data: revenue by source, cost by
    category, and the profit/loss remainder, for the requested trailing
    window (defaults to the last 30 days)."""
    today = date.today()
    start_dt = datetime.combine(today - timedelta(days=days - 1), datetime.min.time())
    end_dt = datetime.combine(today + timedelta(days=1), datetime.min.time())

    room_revenue = _invoice_revenue(db, start_dt, end_dt, bill_type="room")
    mess_revenue = _invoice_revenue(db, start_dt, end_dt, bill_type="mess")
    total_revenue = _invoice_revenue(db, start_dt, end_dt)
    other_revenue = round(total_revenue - room_revenue - mess_revenue, 2)
    procurement_cost, waste_cost = _period_cost(db, start_dt, end_dt)
    profit = total_revenue - procurement_cost - waste_cost

    return {
        "room_revenue": round(room_revenue, 2),
        "mess_revenue": round(mess_revenue, 2),
        "other_revenue": max(other_revenue, 0.0),
        "procurement_cost": round(procurement_cost, 2),
        "waste_cost": round(waste_cost, 2),
        "total_revenue": round(total_revenue, 2),
        "profit": round(profit, 2),
    }


# --- Waste / vendors -------------------------------------------------------

def _waste_by_category_data(db: Session) -> dict:
    month_start = date.today().replace(day=1)
    results = db.query(WasteLog.category, func.sum(WasteLog.quantity), func.sum(WasteLog.cost)).filter(
        WasteLog.created_at >= datetime.combine(month_start, datetime.min.time()),
    ).group_by(WasteLog.category).all()
    return {"labels": [r[0].value if r[0] else "unknown" for r in results],
            "quantities": [float(r[1] or 0) for r in results],
            "costs": [float(r[2] or 0) for r in results]}


@router.get("/waste-by-category")
async def waste_by_category(db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    return _waste_by_category_data(db)


def _vendor_performance_data(db: Session) -> list[dict]:
    """Top vendors by total self-purchase spend in the last 30 days (Daily
    Stock Intake / Smart Intake batches tagged with that vendor) - the only
    real per-vendor signal now that the mess buys directly rather than
    through a PO a vendor could be scored for fulfilling."""
    since = datetime.utcnow() - timedelta(days=30)
    rows = (
        db.query(Vendor.name, func.sum(StockBatch.quantity * StockBatch.unit_cost).label("spend"))
        .join(StockBatch, StockBatch.vendor_id == Vendor.id)
        .filter(Vendor.is_active == True, StockBatch.created_at >= since)
        .group_by(Vendor.id)
        .order_by(func.sum(StockBatch.quantity * StockBatch.unit_cost).desc())
        .limit(10)
        .all()
    )
    return [{"name": name, "spend_30d": round(float(spend or 0), 2)} for name, spend in rows]


@router.get("/vendor-performance")
async def vendor_performance(db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    return _vendor_performance_data(db)


@router.get("/stock-overview")
async def stock_overview(db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    """Read-only stock/procurement summary for the Manager Dashboard widget -
    same underlying figures as /inventory/dashboard, queried directly here
    (rather than calling that router) because /inventory/* is gated on
    inventory:view, which Manager deliberately doesn't hold (that would also
    surface the full Inventory & Procurement nav item - see access.py)."""
    now = datetime.utcnow()
    thirty_days_ago = now - timedelta(days=30)

    active_batches = (
        db.query(StockBatch.item_id, func.sum(StockBatch.quantity).label("total_qty"))
        .filter(StockBatch.is_active == True)
        .group_by(StockBatch.item_id)
        .all()
    )
    qty_by_item = {r.item_id: r.total_qty for r in active_batches}
    last_cost_by_item = {}
    for iid in qty_by_item:
        last_b = (
            db.query(StockBatch.unit_cost)
            .filter(StockBatch.item_id == iid, StockBatch.is_active == True, StockBatch.unit_cost > 0)
            .order_by(StockBatch.created_at.desc())
            .first()
        )
        if last_b:
            last_cost_by_item[iid] = float(last_b.unit_cost)
    inventory_value = sum(qty_by_item.get(iid, 0) * last_cost_by_item.get(iid, 0) for iid in qty_by_item)

    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_procurement = (
        db.query(func.coalesce(func.sum(StockBatch.quantity * StockBatch.unit_cost), 0))
        .filter(StockBatch.created_at >= month_start, StockBatch.is_active == True)
        .scalar()
    )

    items = db.query(InventoryItem).filter(InventoryItem.is_active == True).all()
    low_stock, well_stocked = [], []
    for it in items:
        total_stock = qty_by_item.get(it.id, 0)
        row = {"item_id": it.id, "name": it.name, "unit": it.unit, "total_stock": total_stock, "reorder_level": it.reorder_level}
        if it.reorder_level and it.reorder_level > 0 and total_stock <= it.reorder_level:
            low_stock.append(row)
        elif total_stock > 0:
            well_stocked.append(row)
    low_stock.sort(key=lambda r: r["total_stock"] - r["reorder_level"])
    well_stocked.sort(key=lambda r: r["total_stock"], reverse=True)

    cost_rows = (
        db.query(StockBatch.item_id, InventoryItem.name, func.sum(StockBatch.quantity * StockBatch.unit_cost).label("total_spend"))
        .join(InventoryItem, StockBatch.item_id == InventoryItem.id)
        .filter(StockBatch.created_at >= thirty_days_ago, StockBatch.is_active == True)
        .group_by(StockBatch.item_id, InventoryItem.name)
        .order_by(func.sum(StockBatch.quantity * StockBatch.unit_cost).desc())
        .limit(8)
        .all()
    )
    top_costing_products = [{"item_id": r.item_id, "item_name": r.name, "total_spend": round(float(r.total_spend or 0), 2)} for r in cost_rows]

    return {
        "inventory_value": round(inventory_value, 2),
        "month_procurement": round(float(month_procurement), 2),
        "low_stock_count": len(low_stock),
        "low_stock_items": low_stock[:8],
        "well_stocked_items": well_stocked[:8],
        "top_costing_products": top_costing_products,
    }


# --- Exceptions: things that need a Manager's attention, in one place -----

def _exceptions_data(db: Session, days: int) -> dict:
    today = date.today()
    start_dt = datetime.combine(today - timedelta(days=days - 1), datetime.min.time())
    end_dt = datetime.combine(today + timedelta(days=1), datetime.min.time())

    discounted = db.query(Invoice).filter(
        Invoice.created_at >= start_dt, Invoice.created_at < end_dt, Invoice.discount > 0,
    ).all()
    voided = db.query(Invoice).filter(
        Invoice.created_at >= start_dt, Invoice.created_at < end_dt, Invoice.status == InvoiceStatus.VOID,
    ).all()

    corrections_pending = db.query(InvoiceEditRequest).filter(InvoiceEditRequest.status == EditRequestStatus.PENDING).count()
    corrections_period = db.query(InvoiceEditRequest).filter(
        InvoiceEditRequest.requested_at >= start_dt, InvoiceEditRequest.requested_at < end_dt,
    ).count()
    menu_changes_pending = db.query(MenuItemEditRequest).filter(MenuItemEditRequest.status == EditRequestStatus.PENDING).count()

    # Snapshot (right now), not period-bound - "how many guests are overdue
    # to leave" doesn't have a meaningful start/end, only a current count.
    overdue_departures = db.query(Booking).filter(
        Booking.check_out < today, Booking.status == "checked_in",
        or_(Booking.nature_of_duty.is_(None), Booking.nature_of_duty != "hra"),
    ).count()

    overdue_invoices = [
        inv for inv in db.query(Invoice).filter(Invoice.status == InvoiceStatus.ISSUED, Invoice.due_date < today).all()
        if float(inv.total_amount) - float(inv.amount_paid or 0) > 0.01
    ]

    anomaly_alerts = db.query(Alert).filter(
        Alert.module == "procurement", Alert.entity_type.in_(["spend_zscore", "benford_line_items"]),
        Alert.status.in_([AlertStatus.NEW, AlertStatus.ACKNOWLEDGED]),
    ).order_by(Alert.created_at.desc()).limit(10).all()

    return {
        "period_days": days,
        "discounted_count": len(discounted),
        "discounted_amount": round(sum(float(i.discount) for i in discounted), 2),
        "void_count": len(voided),
        "void_amount": round(sum(float(i.total_amount) for i in voided), 2),
        "corrections_pending": corrections_pending,
        "corrections_period": corrections_period,
        "menu_changes_pending": menu_changes_pending,
        "overdue_departures": overdue_departures,
        "overdue_invoices_count": len(overdue_invoices),
        "overdue_invoices_amount": round(sum(float(i.total_amount) - float(i.amount_paid or 0) for i in overdue_invoices), 2),
        "anomaly_alerts": [
            {"id": a.id, "title": a.title, "message": a.message, "severity": a.severity.value, "created_at": a.created_at}
            for a in anomaly_alerts
        ],
    }


@router.get("/exceptions")
async def exceptions_report(days: int = Query(7, ge=1, le=90), db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    """Things that need a Manager's attention and would otherwise require
    going and looking in five different places: discounts, voids, pending
    bill/menu corrections, overdue departures/invoices, and open procurement
    anomaly alerts."""
    return _exceptions_data(db, days)


# --- Staff activity summary ------------------------------------------------

def _audit_summary_data(db: Session, days: int) -> dict:
    start_dt = datetime.combine(date.today() - timedelta(days=days - 1), datetime.min.time())
    total = db.query(AuditLog).filter(AuditLog.timestamp >= start_dt).count()
    by_user = (
        db.query(AuditLog.user_name, func.count(AuditLog.id))
        .filter(AuditLog.timestamp >= start_dt)
        .group_by(AuditLog.user_name)
        .order_by(func.count(AuditLog.id).desc())
        .limit(10)
        .all()
    )
    by_action = (
        db.query(AuditLog.action, func.count(AuditLog.id))
        .filter(AuditLog.timestamp >= start_dt)
        .group_by(AuditLog.action)
        .order_by(func.count(AuditLog.id).desc())
        .all()
    )
    return {
        "period_days": days,
        "total_actions": total,
        "by_user": [{"user_name": u or "System", "count": c} for u, c in by_user],
        "by_action": [{"action": a.value, "count": c} for a, c in by_action],
    }


@router.get("/audit-summary")
async def audit_summary(days: int = Query(7, ge=1, le=90), db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    """'Who did what' rollup over Audit Log - the raw log stays available
    for a full drill-down, but a Manager checking in doesn't have to read
    every row to get a sense of activity volume."""
    return _audit_summary_data(db, days)


# --- Export ----------------------------------------------------------------

@router.get("/export")
async def export_report(days: int = Query(30, ge=7, le=365), db: Session = Depends(get_db), current_user=Depends(PermissionChecker("reports", "view"))):
    """One workbook covering everything on the Reports page for the
    selected window - a Manager reporting to an oversight board previously
    had nothing to hand over from this page at all."""
    kpis = _dashboard_kpis(db)
    revenue = _revenue_detail_data(db, days)
    occupancy = _occupancy_trend_data(db, days)
    waste = _waste_by_category_data(db)
    vendors = _vendor_performance_data(db)
    exceptions = _exceptions_data(db, min(days, 90))
    audit = _audit_summary_data(db, min(days, 90))

    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Summary"
    ws.append(["EME MESS - Reports Summary", f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"])
    ws["A1"].font = bold
    ws.append([])
    ws.append(["This Month at a Glance"])
    ws["A3"].font = bold
    for label, value in [
        ("Today's revenue", kpis["today_revenue"]), ("Occupancy rate (%)", kpis["occupancy_rate"]),
        ("Total stock value", kpis["total_stock_value"]), ("Waste cost this month", kpis["waste_cost_month"]),
        ("Open alerts", kpis["open_alerts"]), ("Guests today", kpis["total_guests_today"]),
        ("Low stock items", kpis["low_stock_count"]), ("Open security incidents", kpis["open_incidents"]),
        ("Mess revenue this month", kpis["mess_revenue_month"]), ("Unpaid mess bills", kpis["unpaid_mess_bills"]),
        ("Active members", kpis["active_member_count"]), ("Outstanding balance", kpis["outstanding_balance"]),
        ("Unsettled invoices", kpis["unsettled_invoice_count"]), ("Discounts this month", kpis["discounts_month"]),
    ]:
        ws.append([label, value])

    ws2 = wb.create_sheet("Revenue & Cost")
    ws2.append([f"Revenue, cost and profit - last {days} days"])
    ws2["A1"].font = bold
    ws2.append(["Date", "Revenue", "Cost", "Profit"])
    for i, label in enumerate(revenue["labels"]):
        ws2.append([label, revenue["revenue"][i], revenue["cost"][i], revenue["profit"][i]])
    ws2.append([])
    ws2.append(["Room revenue total", revenue["room_revenue_total"]])
    ws2.append(["Mess revenue total", revenue["mess_revenue_total"]])
    ws2.append(["Other revenue total", revenue["other_revenue_total"]])
    ws2.append(["Total revenue", revenue["revenue_total"], f"vs previous period: {revenue['revenue_pct_change']}%"])
    ws2.append(["Total cost", revenue["cost_total"], f"vs previous period: {revenue['cost_pct_change']}%"])
    ws2.append(["Total profit", revenue["profit_total"], f"vs previous period: {revenue['profit_pct_change']}%"])

    ws3 = wb.create_sheet("Occupancy")
    ws3.append([f"Occupancy rate - last {days} days"])
    ws3["A1"].font = bold
    ws3.append(["Date", "Occupancy %"])
    for i, label in enumerate(occupancy["labels"]):
        ws3.append([label, occupancy["values"][i]])
    ws3.append([])
    ws3.append(["Average this period", occupancy["avg_current"]])
    ws3.append(["Average previous period", occupancy["avg_previous"]])
    ws3.append(["Change", f"{occupancy['pct_change']}%"])

    ws4 = wb.create_sheet("Waste & Vendors")
    ws4.append(["Waste by Category (this month)"])
    ws4["A1"].font = bold
    ws4.append(["Category", "Quantity", "Cost"])
    for i, label in enumerate(waste["labels"]):
        ws4.append([label, waste["quantities"][i], waste["costs"][i]])
    ws4.append([])
    ws4.append(["Top Vendors by Spend (30 days)"])
    ws4[f"A{ws4.max_row}"].font = bold
    ws4.append(["Vendor", "Spend"])
    for v in vendors:
        ws4.append([v["name"], v["spend_30d"]])

    ws5 = wb.create_sheet("Exceptions")
    ws5.append([f"Exceptions - last {exceptions['period_days']} days"])
    ws5["A1"].font = bold
    for label, value in [
        ("Bills discounted (count)", exceptions["discounted_count"]),
        ("Bills discounted (amount)", exceptions["discounted_amount"]),
        ("Bills voided (count)", exceptions["void_count"]),
        ("Bills voided (amount)", exceptions["void_amount"]),
        ("Bill corrections pending", exceptions["corrections_pending"]),
        ("Bill corrections requested this period", exceptions["corrections_period"]),
        ("Menu changes pending", exceptions["menu_changes_pending"]),
        ("Guests overdue to leave (now)", exceptions["overdue_departures"]),
        ("Overdue invoices (count, now)", exceptions["overdue_invoices_count"]),
        ("Overdue invoices (amount, now)", exceptions["overdue_invoices_amount"]),
    ]:
        ws5.append([label, value])
    ws5.append([])
    ws5.append(["Open Procurement Anomaly Alerts"])
    ws5[f"A{ws5.max_row}"].font = bold
    ws5.append(["Title", "Severity", "Created"])
    for a in exceptions["anomaly_alerts"]:
        ws5.append([a["title"], a["severity"], a["created_at"].strftime("%Y-%m-%d") if a["created_at"] else ""])

    ws6 = wb.create_sheet("Staff Activity")
    ws6.append([f"Staff activity - last {audit['period_days']} days ({audit['total_actions']} total actions)"])
    ws6["A1"].font = bold
    ws6.append(["User", "Actions"])
    for row in audit["by_user"]:
        ws6.append([row["user_name"], row["count"]])
    ws6.append([])
    ws6.append(["Action type", "Count"])
    for row in audit["by_action"]:
        ws6.append([row["action"], row["count"]])

    for sheet in wb.worksheets:
        for col_idx in range(1, 5):
            sheet.column_dimensions[chr(64 + col_idx)].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reports_summary_{date.today().isoformat()}.xlsx"},
    )
