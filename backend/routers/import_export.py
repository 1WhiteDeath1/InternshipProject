"""Excel import/export router.

Permission model: every action here (template download, import, export) is
gated uniformly on the `import_export` permission - NOT on the target data
module's own permission (inventory/vendors/rooms/bookings). That used to be
the case, but "vendors"/"rooms"/"opening_stock" were never real RBAC modules
anywhere in access.py, so those cards silently 403'd for every role,
including Manager (the only role that even holds import_export and can see
this page's nav item at all). Gating on import_export throughout means: if
you can see this page, every card on it actually works."""
import io
import uuid
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from backend.database import get_db
from backend.models import (
    InventoryItem, InventoryCategory, Vendor, Room,
    Booking, BookingStatus, StockBatch,
)
from backend.auth import get_current_user, check_permission
from backend.audit import log_audit, AuditAction
from backend.logging_config import get_logger

logger = get_logger("app")
router = APIRouter()


TEMPLATES = {
    "inventory": ["sku", "name", "category", "unit", "reorder_level", "reorder_quantity"],
    "vendors": ["name", "contact_person", "phone", "email", "address", "tax_id", "payment_terms"],
    "rooms": ["room_number", "room_type", "floor", "capacity", "base_price", "amenities"],
    "bookings": ["guest_name", "guest_phone", "guest_email", "room_number", "check_in", "check_out", "adults", "children"],
    "opening_stock": ["sku", "batch_number", "quantity", "bin_location", "expiry_date", "unit_cost"],
}


def _parse_date(value) -> date:
    """openpyxl hands back a datetime/date object for a real Excel date cell,
    or a plain string if the column was typed as text - accept either."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()


@router.get("/template/{module}")
async def get_template(module: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "import_export", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    if module not in TEMPLATES:
        raise HTTPException(status_code=400, detail="Unknown module")

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(TEMPLATES[module])
    ws.append(["Example"] * len(TEMPLATES[module]))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename={module}_template.xlsx"})


@router.post("/import/{module}")
async def import_data(module: str, file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "import_export", "create"):
        raise HTTPException(status_code=403, detail="Permission denied")
    if module not in TEMPLATES:
        raise HTTPException(status_code=400, detail="Unknown module")

    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    errors = []
    imported = 0
    skipped = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = dict(zip(headers, row))
            if module == "inventory":
                cat = db.query(InventoryCategory).filter(InventoryCategory.name == data.get("category", "General")).first()
                if not cat:
                    cat = InventoryCategory(name=data.get("category", "General"))
                    db.add(cat)
                    db.commit()
                    db.refresh(cat)
                if db.query(InventoryItem).filter(InventoryItem.sku == data["sku"]).first():
                    skipped += 1
                    continue
                item = InventoryItem(sku=data["sku"], name=data["name"], category_id=cat.id, unit=data.get("unit", "pcs"),
                                    reorder_level=float(data.get("reorder_level", 0) or 0), reorder_quantity=float(data.get("reorder_quantity", 0) or 0))
                db.add(item)
                imported += 1

            elif module == "vendors":
                if db.query(Vendor).filter(Vendor.name == data["name"]).first():
                    skipped += 1
                    continue
                v = Vendor(**{k: v for k, v in data.items() if v})
                db.add(v)
                imported += 1

            elif module == "rooms":
                if db.query(Room).filter(Room.room_number == data["room_number"]).first():
                    skipped += 1
                    continue
                r = Room(room_number=data["room_number"], room_type=data.get("room_type", "standard"), floor=int(data.get("floor", 1)),
                        capacity=int(data.get("capacity", 2)), base_price=float(data["base_price"]), amenities=data.get("amenities"))
                db.add(r)
                imported += 1

            elif module == "bookings":
                room = db.query(Room).filter(Room.room_number == data["room_number"]).first()
                if not room:
                    raise ValueError(f"Room {data.get('room_number')} not found")
                check_in = _parse_date(data["check_in"])
                check_out = _parse_date(data["check_out"])
                # Natural-key dedup: same guest, same room, same arrival date
                # already on file - re-running the same sheet shouldn't
                # duplicate bookings.
                if db.query(Booking).filter(Booking.guest_name == data["guest_name"], Booking.room_id == room.id, Booking.check_in == check_in).first():
                    skipped += 1
                    continue
                booking = Booking(
                    booking_reference=f"TMP-{uuid.uuid4().hex}", guest_name=data["guest_name"],
                    guest_phone=data.get("guest_phone"), guest_email=data.get("guest_email"),
                    room_id=room.id, check_in=check_in, check_out=check_out,
                    adults=int(data.get("adults") or 1), children=int(data.get("children") or 0),
                    status=BookingStatus.CONFIRMED, processed_by=current_user.id,
                )
                db.add(booking)
                db.commit()
                db.refresh(booking)
                # Bulk-imported records land as CONFIRMED, not CHECKED_IN - a
                # real check-in (attendant assignment, room-readiness gate)
                # still has to happen through the normal Bookings flow;
                # importing is for getting historical/paper records into the
                # system, not for skipping that gate.
                booking.booking_reference = f"BK-{datetime.utcnow().strftime('%Y%m%d')}-{booking.id:04d}"
                db.commit()
                imported += 1

            elif module == "opening_stock":
                item = db.query(InventoryItem).filter(InventoryItem.sku == data["sku"]).first()
                if not item:
                    raise ValueError(f"Inventory item with SKU {data.get('sku')} not found - import Inventory first")
                if db.query(StockBatch).filter(StockBatch.item_id == item.id, StockBatch.batch_number == data["batch_number"]).first():
                    skipped += 1
                    continue
                batch = StockBatch(
                    item_id=item.id, batch_number=data["batch_number"], quantity=float(data.get("quantity", 0) or 0),
                    bin_location=data.get("bin_location"),
                    expiry_date=_parse_date(data["expiry_date"]) if data.get("expiry_date") else None,
                    unit_cost=float(data.get("unit_cost", 0) or 0),
                )
                db.add(batch)
                imported += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    db.commit()
    log_audit(db, current_user.id, current_user.full_name, AuditAction.IMPORT, module,
              reason=f"Imported {imported}, skipped {skipped}, {len(errors)} errors",
              after_state={"imported": imported, "skipped": skipped, "errors": len(errors)})
    return {
        "imported": imported, "skipped": skipped, "errors": len(errors), "error_details": errors[:10],
        "message": f"Imported {imported} new row(s), skipped {skipped} existing row(s), {len(errors)} error(s)",
    }


@router.get("/export/{module}")
async def export_data(module: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "import_export", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    wb = Workbook()
    ws = wb.active

    if module == "inventory":
        ws.append(["ID", "SKU", "Name", "Category", "Unit", "Reorder Level", "Reorder Qty", "Total Stock"])
        items = db.query(InventoryItem).filter(InventoryItem.is_active == True).all()
        for item in items:
            batches = db.query(StockBatch).filter(StockBatch.item_id == item.id, StockBatch.is_active == True).all()
            total = sum(b.quantity for b in batches)
            ws.append([item.id, item.sku, item.name, item.category.name if item.category else "", item.unit, item.reorder_level, item.reorder_quantity, total])
    elif module == "bookings":
        ws.append(["ID", "Reference", "Guest", "Phone", "Room", "Check In", "Check Out", "Status", "Total"])
        bookings = db.query(Booking).order_by(Booking.created_at.desc()).limit(1000).all()
        for b in bookings:
            ws.append([b.id, b.booking_reference, b.guest_name, b.guest_phone, b.room.room_number if b.room else "", b.check_in, b.check_out, b.status.value, float(b.total_amount) if b.total_amount else ""])
    elif module == "vendors":
        ws.append(["ID", "Name", "Contact", "Phone", "Email", "Address", "Tax ID"])
        vendors = db.query(Vendor).filter(Vendor.is_active == True).all()
        for v in vendors:
            ws.append([v.id, v.name, v.contact_person, v.phone, v.email, v.address, v.tax_id])
    elif module == "rooms":
        ws.append(["ID", "Room Number", "Type", "Floor", "Capacity", "Base Price", "Amenities", "Status"])
        rooms = db.query(Room).filter(Room.is_active == True).order_by(Room.room_number).all()
        for r in rooms:
            ws.append([r.id, r.room_number, getattr(r.room_type, "value", r.room_type), r.floor, r.capacity, float(r.base_price), r.amenities, getattr(r.status, "value", r.status)])
    elif module == "opening_stock":
        ws.append(["ID", "SKU", "Item Name", "Batch Number", "Quantity", "Bin Location", "Expiry Date", "Unit Cost"])
        batches = db.query(StockBatch).filter(StockBatch.is_active == True).order_by(StockBatch.received_date.desc()).all()
        for b in batches:
            ws.append([b.id, b.item.sku if b.item else "", b.item.name if b.item else "", b.batch_number, b.quantity, b.bin_location, b.expiry_date, float(b.unit_cost or 0)])
    elif module == "audit":
        from sqlalchemy import desc
        from backend.models import AuditLog
        ws.append(["ID", "User", "Action", "Entity", "Entity ID", "Reason", "Department", "Timestamp", "IP"])
        logs = db.query(AuditLog).order_by(desc(AuditLog.timestamp)).limit(5000).all()
        for l in logs:
            ws.append([l.id, l.user_name, l.action.value if l.action else "", l.entity_type, l.entity_id, l.reason, l.department, l.timestamp, l.ip_address])
    else:
        raise HTTPException(status_code=400, detail="Export module not supported")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename={module}_export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"})
