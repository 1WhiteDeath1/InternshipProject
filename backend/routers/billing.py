"""Billing and invoicing router."""
import hashlib
import io
import json
import uuid
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict
from fastapi import APIRouter, Depends, HTTPException, Query, Request, Body
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import or_, func
from sqlalchemy.orm import Session, joinedload
from backend.database import get_db
from backend.models import (
    Invoice, InvoiceItem, InvoicePayment, Booking, BookingCharge, InvoiceStatus,
    ClientCategory, MealAttendance, KitchenOrder, User, Guest, AlertSeverity,
    InvoiceEditRequest, EditRequestStatus, InventoryItem, StockBatch, WasteLog, GasChargeRate,
)
from backend.schemas import (
    InvoiceItemCreate, BookingChargeCreate, PaymentCreate,
    InvoiceEditRequestCreate, InvoiceEditDecision,
    MasterBillLineCreate, MasterBillLineCorrection, LastDebitBalanceUpdate,
)
from backend.auth import get_current_user, check_permission
from backend.audit import log_audit, serialize_model, AuditAction
from backend.alerts import create_alert
from backend.logging_config import get_logger
from backend.services.mess_billing_calc import get_setting_float, get_setting_str
from backend.services.room_pricing import reprice_for_departure
from backend.services.mess_charge_calc import (
    compute_unbilled_mess_total, compute_unbilled_gas_total, meal_multiplier_for_booking as _meal_multiplier,
    walkin_meal_multiplier as _walkin_meal_multiplier,
)
from backend.services.master_bill import assemble_rows, MASTER_BILL_PRESET_HEADS

logger = get_logger("app")
router = APIRouter()

# Split-payment methods (Workflow 1's Payment Summary Breakdown groups by
# exactly these three). Voucher Number is mandatory for Online/AG Branch,
# optional for Bank Transfer, unused for Cash Deposit - enforced in
# PaymentCreate's validator (schemas/billing.py).
PAYMENT_METHOD_ONLINE = "Online/AG Branch"
PAYMENT_METHOD_BANK = "Bank Transfer"
PAYMENT_METHOD_CASH = "Cash Deposit"
PAYMENT_METHODS = (PAYMENT_METHOD_ONLINE, PAYMENT_METHOD_BANK, PAYMENT_METHOD_CASH)


def _prior_outstanding_for_guest(db: Session, guest_identity_id: int) -> float:
    """Sum of everything still owed across every PRIOR bill tied to this
    persistent Guest identity (matched by CNIC/phone at check-in - see
    models/guests.py) - covers both a past room stay (via Booking.guest_id)
    and a past standalone walk-in mess visit (Invoice.guest_id directly).
    Used to auto-fill a brand-new bill's Last Debit Balance; the Clerk can
    still edit it afterward via set_last_debit_balance."""
    prior = db.query(Invoice).outerjoin(Booking, Invoice.booking_id == Booking.id).filter(
        Invoice.status != InvoiceStatus.VOID,
        or_(Invoice.guest_id == guest_identity_id, Booking.guest_id == guest_identity_id),
    ).all()
    total = 0.0
    for inv in prior:
        owed = _invoice_net_owed(inv) - float(inv.amount_paid or 0)
        if owed > 0:
            total += owed
    return round(total, 2)


def _invoice_net_owed(inv: Invoice) -> float:
    """What the guest actually owes on this bill, including any carried-
    forward balance - the same figure the Master Bill's Net Debits row
    represents (see master_bill.assemble_rows), used to decide
    draft->issued->partially_paid->paid transitions."""
    return float(inv.total_amount or 0) + float(inv.last_debit_balance or 0)


def _build_payment_breakdown(payments: list) -> dict:
    """Groups a bill's payment rows into the Payment Summary Breakdown
    Workflow 1 wants on the Master Bill: one bucket per method plus an
    "other" catch-all for any payment recorded before this split-payment
    architecture existed (free-text method that isn't one of the three)."""
    by_method = {PAYMENT_METHOD_ONLINE: 0.0, PAYMENT_METHOD_BANK: 0.0, PAYMENT_METHOD_CASH: 0.0, "other": 0.0}
    vouchers = []
    for p in payments:
        amount = float(p.amount)
        if p.method in by_method:
            by_method[p.method] += amount
        else:
            by_method["other"] += amount
        if p.voucher_number:
            vouchers.append(p.voucher_number)
    total_paid = sum(by_method.values())
    return {
        "online_ag_branch": round(by_method[PAYMENT_METHOD_ONLINE], 2),
        "online_ag_branch_vouchers": vouchers,
        "bank_transfer": round(by_method[PAYMENT_METHOD_BANK], 2),
        "cash_deposit": round(by_method[PAYMENT_METHOD_CASH], 2),
        "other": round(by_method["other"], 2),
        "total_paid": round(total_paid, 2),
        "payments": [{"id": p.id, "amount": float(p.amount), "method": p.method, "voucher_number": p.voucher_number,
                      "notes": p.notes, "created_at": p.created_at} for p in payments],
    }


def _apply_invoice_payment_status(inv: Invoice) -> None:
    net_owed = _invoice_net_owed(inv)
    paid = float(inv.amount_paid or 0)
    if paid >= net_owed - 0.01:
        inv.status = InvoiceStatus.PAID
    elif paid > 0.01:
        inv.status = InvoiceStatus.PARTIALLY_PAID
    elif inv.status == InvoiceStatus.DRAFT:
        inv.status = InvoiceStatus.ISSUED


def _build_invoice(db: Session, booking: Booking, items: list, current_user, issue_date: date, due_date: date, tax_amount: float = 0.0, discount: float = 0.0, notes: str = None, bill_type: str = "combined", *, guest: Guest = None) -> Invoice:
    """Shared invoice-assembly logic used by Instant Checkout (room stay) and
    the walk-in mess bill: assigns a race-free invoice number and creates the
    InvoiceItem rows. Owner is a Booking (booking passed) or a standalone
    walk-in Guest (guest= passed) - exactly one. Callers are responsible for
    pricing items at whatever the guest actually owes (e.g. the consolidated
    Extra Messing/Gas totals already apply the meal multiplier themselves -
    see mess_charge_calc.compute_unbilled_mess_total)."""
    owner = {"booking_id": booking.id, "guest_id": None} if booking is not None else {"booking_id": None, "guest_id": guest.id}
    total = sum(item.unit_price * item.quantity for item in items)

    # Auto-fill Last Debit Balance from this same persistent Guest's other
    # still-unpaid bills (a prior stay or walk-in visit) - only when the
    # identity is actually known (guest_id set at check-in/registration);
    # the Clerk can still override it afterward via set_last_debit_balance.
    guest_identity_id = booking.guest_id if booking is not None else guest.id
    auto_last_debit_balance = _prior_outstanding_for_guest(db, guest_identity_id) if guest_identity_id else 0.0

    invoice = Invoice(
        invoice_number=f"TMP-{uuid.uuid4().hex}", **owner,
        issue_date=issue_date, due_date=due_date,
        subtotal=total, tax_amount=tax_amount,
        discount=discount, total_amount=total + tax_amount - discount,
        last_debit_balance=auto_last_debit_balance,
        notes=notes, created_by=current_user.id, bill_type=bill_type,
        # A checkout bill handed to the guest is final the moment it exists -
        # 'issued', not 'draft', so revenue stats and overdue tracking see it.
        status=InvoiceStatus.ISSUED,
    )
    db.add(invoice)
    db.commit()
    db.refresh(invoice)

    # Friendly number derived from the autoincrement id (assigned above), not a
    # pre-insert count() - avoids a race where two concurrent creates generate the
    # same invoice number.
    invoice.invoice_number = f"INV-{datetime.utcnow().strftime('%Y%m')}-{invoice.id:05d}"

    for item in items:
        ii = InvoiceItem(
            invoice_id=invoice.id, description=item.description,
            quantity=item.quantity, unit_price=item.unit_price,
            total_price=item.unit_price * item.quantity,
        )
        db.add(ii)
    db.commit()
    db.refresh(invoice)
    return invoice


@router.get("/invoices")
async def list_invoices(
    status: str = "", search: str = "", bill_type: str = "",
    page: int = Query(1, ge=1), page_size: int = Query(25, ge=1),
    db: Session = Depends(get_db), current_user=Depends(get_current_user),
):
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    # Outer-joined (not filtered) so search can reach into the owning
    # Booking/Guest - each Invoice owns at most one of the two (owner is
    # exactly one per _build_invoice), so this can't fan out duplicate rows.
    query = db.query(Invoice).outerjoin(Booking, Invoice.booking_id == Booking.id).outerjoin(Guest, Invoice.guest_id == Guest.id)
    if status:
        query = query.filter(Invoice.status == status)
    if bill_type:
        query = query.filter(Invoice.bill_type == bill_type)
    if search:
        # Customer Name / Rank / Service Number / Voucher Number / Bill
        # Serial Number - Booking carries rank/pa_number/online_voucher_no
        # per-stay; a standalone walk-in Guest (mess-only invoice with no
        # booking) only has full_name/id_number, so both sides are covered.
        query = query.filter(or_(
            Invoice.invoice_number.contains(search),
            Invoice.bill_serial_number.contains(search),
            Booking.guest_name.contains(search),
            Booking.rank.contains(search),
            Booking.pa_number.contains(search),
            Booking.online_voucher_no.contains(search),
            Guest.full_name.contains(search),
            Guest.id_number.contains(search),
        ))

    total = query.count()
    invoices = query.order_by(Invoice.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return {"items": [
        {"id": inv.id, "invoice_number": inv.invoice_number, "booking_id": inv.booking_id,
         "customer_guest_id": (inv.booking.guest_id if inv.booking else None) or inv.guest_id,
         "guest_name": inv.booking.guest_name if inv.booking else (inv.guest.full_name if inv.guest else None),
         "room_number": inv.booking.room.room_number if inv.booking and inv.booking.room else None,
         "issue_date": inv.issue_date, "due_date": inv.due_date,
         "subtotal": float(inv.subtotal), "tax_amount": float(inv.tax_amount),
         "discount": float(inv.discount), "total_amount": float(inv.total_amount),
         "amount_paid": float(inv.amount_paid), "status": inv.status.value,
         "bill_type": inv.bill_type or "combined", "is_complimentary": bool(inv.is_complimentary),
         "notes": inv.notes, "created_at": inv.created_at,
         "items": [{"id": i.id, "description": i.description, "quantity": i.quantity,
                    "unit_price": float(i.unit_price), "total_price": float(i.total_price)} for i in inv.items]} for inv in invoices], "total": total}


@router.get("/customers/{guest_id}/history")
async def customer_history(guest_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Workflow 2's Customer Summary Modal - the persistent Guest identity
    plus every Master Bill (Invoice) and payment/receipt ever tied to it,
    across every stay/visit, not just the one the Clerk searched from."""
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    guest = db.query(Guest).filter(Guest.id == guest_id).first()
    if not guest:
        raise HTTPException(status_code=404, detail="Customer not found")

    invoices = db.query(Invoice).outerjoin(Booking, Invoice.booking_id == Booking.id).filter(
        or_(Invoice.guest_id == guest_id, Booking.guest_id == guest_id),
    ).order_by(Invoice.created_at.desc()).all()

    bills = []
    for inv in invoices:
        payments = db.query(InvoicePayment).filter(InvoicePayment.invoice_id == inv.id).order_by(InvoicePayment.created_at.desc()).all()
        bills.append({
            "id": inv.id, "invoice_number": inv.invoice_number, "bill_type": inv.bill_type or "combined",
            "bill_serial_number": inv.bill_serial_number, "issue_date": inv.issue_date, "status": inv.status.value,
            "total_amount": float(inv.total_amount), "last_debit_balance": float(inv.last_debit_balance or 0),
            "amount_paid": float(inv.amount_paid or 0),
            "room_number": inv.booking.room.room_number if inv.booking and inv.booking.room else None,
            "payments": [{"id": p.id, "amount": float(p.amount), "method": p.method, "voucher_number": p.voucher_number,
                          "created_at": p.created_at} for p in payments],
        })

    return {
        "guest": {"id": guest.id, "full_name": guest.full_name, "phone": guest.phone,
                  "id_type": guest.id_type, "id_number": guest.id_number, "unit_address": guest.unit_address},
        "bills": bills,
    }


def _gas_percentage(db: Session) -> float:
    row = db.query(GasChargeRate).first()
    return float(row.percentage) if row else 0.0


# Display labels for compute_booking_price()'s components dict keys (see
# room_pricing.py) - covers every pricing mode's component set: rate-card/
# member-rate (rent/electricity/generator/gas/internet), tariff matrix
# (tariff_rate), official duty (da_charge), and HRA monthly (hra_rank_rate,
# utility_charge - rarely hit here since HRA residents settle via the
# monthly Mess Bill instead, see instant_checkout's guard below).
ROOM_COMPONENT_LABELS = {
    "rent": "Room Rent", "electricity": "Electricity Charges", "generator": "Generator Charges",
    "gas": "Sui Gas Bill", "internet": "Internet Charges", "tariff_rate": "Tariff Rate",
    "da_charge": "Daily Allowance (DA) Charge", "hra_rank_rate": "HRA Rank Rate", "utility_charge": "Room Utilities",
}
# Components that are already an absolute monthly figure, not a nightly rate
# to multiply by nights (see compute_booking_price's hra_monthly branch).
_MONTHLY_ROOM_COMPONENTS = {"hra_rank_rate", "utility_charge"}


def _gather_unbilled_items(db: Session, booking_id: int, *, mess_override: float = None, gas_override: float = None, room_component_overrides: Dict[str, float] = None):
    """Collects everything not yet invoiced for a booking, split into the two
    checkout bills: room_items (itemized rate components - rent/electricity/
    etc, see ROOM_COMPONENT_LABELS - plus non-mess ad-hoc charges like Dhobi/
    Breakage/Allied) and mess_items. Extra Messing is one consolidated line =
    everything ordered through the kitchen (routine meals + a la carte, see
    mess_charge_calc.compute_unbilled_mess_total); Sui Gas Charges on Messing
    is that total x the Kitchen NCO's gas percentage. All of these can be
    overridden by the caller (Clerk Desk's checkout preview). Any other
    ad-hoc mess-tagged BookingCharge (logged via "Custom...") stays its own
    separate line. Returned as a dict so both the read-only balance check and
    the mutating checkout can share it - every item carries a component_key
    (computed/correctable rate lines) or charge_id (ad-hoc BookingCharge,
    removable) so the running-balance preview can tell them apart from a
    fixed structural line (Extra Mattress/Late Fee) with neither set."""
    booking = db.query(Booking).filter(Booking.id == booking_id).first()
    room_items, mess_items, unpriced = [], [], []
    amounts = {"room": 0.0, "mess_total": 0.0, "gas_total": 0.0,
               "room_charges": 0.0, "other_mess_charges": 0.0}
    # Preview must reflect the same meal scaling the invoice will apply, or the
    # Clerk Desk total shown before checkout won't match the invoice produced.
    meal_multiplier = _meal_multiplier(db, booking)

    room_total = float(booking.total_amount) if booking.total_amount else 0.0
    late_fee = float(booking.late_checkout_fee) if booking.late_checkout_fee else 0.0
    pricing_snapshot = {}
    try:
        pricing_snapshot = json.loads(booking.rate_breakdown or "{}") or {}
    except (ValueError, TypeError):
        pass
    mattress_total = float(pricing_snapshot.get("mattress_total") or 0)
    check_out_label = booking.check_out
    # For a guest still in-house, preview the ACTUAL-stay price as of today
    # (overstayed nights added, unused nights dropped) - the same re-price
    # perform_check_out will apply, so preview always equals invoice. A
    # checked-out booking's stored figures already reflect it.
    if booking.status.value == "checked_in":
        effective_out, projected = reprice_for_departure(db, booking, date.today())
        if projected is not None:
            room_total = projected["total"]
            mattress_total = float(projected.get("mattress_total") or 0)
            check_out_label = effective_out
            pricing_snapshot = projected

    if room_total:
        # Break the paper bill's separate heads out of the booking total:
        # Extra Mattress (from the pricing snapshot) and the late checkout fee
        # each get their own line, and the rate itself is split into its real
        # components (rent/electricity/generator/gas/internet, etc, from the
        # pricing snapshot) rather than one lump "Guest Room Charges" figure -
        # each individually overridable via room_component_overrides. Extra
        # Mattress/Late Fee/any ad-hoc BookingCharge lines below stay
        # itemized exactly as computed, untouched by component overrides.
        base_room = max(room_total - late_fee - mattress_total, 0.0)
        nights = pricing_snapshot.get("nights") or 1
        components = pricing_snapshot.get("components") or {}
        overrides = room_component_overrides or {}
        room_no = booking.room.room_number if booking.room else ""
        date_range = f"(From {booking.check_in.strftime('%d-%m-%y')} to {check_out_label.strftime('%d-%m-%y')})"
        if components:
            base_room = 0.0
            for key, rate_value in components.items():
                computed = float(rate_value) if key in _MONTHLY_ROOM_COMPONENTS else round(float(rate_value) * nights, 2)
                line_total = overrides.get(key, computed)
                base_room += line_total
                label = f"{ROOM_COMPONENT_LABELS.get(key, key.replace('_', ' ').title())} - Room {room_no} {date_range}".strip()
                room_items.append(InvoiceItemCreate(description=label, quantity=1, unit_price=line_total, component_key=key))
        else:
            # No component snapshot available (legacy booking pre-dating this
            # breakdown, or a mode with no components computed) - fall back
            # to one overridable lump line, same as before this change.
            base_room = overrides.get("base", base_room)
            label = f"Guest Room Charges - Room {room_no} {date_range}".strip()
            room_items.append(InvoiceItemCreate(description=label, quantity=1, unit_price=base_room, component_key="base"))
        room_total = base_room + late_fee + mattress_total
        amounts["room"] = room_total
        if mattress_total > 0:
            room_items.append(InvoiceItemCreate(description="Extra Mattress", quantity=1, unit_price=mattress_total))
        if late_fee > 0:
            room_items.append(InvoiceItemCreate(description="Late Checkout Fee", quantity=1, unit_price=late_fee))

    mess_total, mess_unpriced, attendance_rows, ala_carte_orders = compute_unbilled_mess_total(
        db, booking_id=booking_id, meal_multiplier=meal_multiplier,
    )
    unpriced.extend(mess_unpriced)
    mess_total = mess_override if mess_override is not None else mess_total
    if mess_total > 0:
        # Multiplier already baked in by compute_unbilled_mess_total - the
        # price on this line is exactly what the invoice will charge.
        mess_items.append(InvoiceItemCreate(description="Extra Messing", quantity=1, unit_price=mess_total, component_key="mess_total"))
    amounts["mess_total"] = mess_total

    gas_total = gas_override if gas_override is not None else compute_unbilled_gas_total(
        db, attendance_rows, meal_multiplier=meal_multiplier, fallback_percentage=_gas_percentage(db),
    )
    if gas_total > 0:
        mess_items.append(InvoiceItemCreate(description="Sui Gas Charges on Messing", quantity=1, unit_price=gas_total, component_key="gas_total"))
    amounts["gas_total"] = gas_total

    charge_rows = db.query(BookingCharge).filter(
        BookingCharge.booking_id == booking_id, BookingCharge.invoiced_at.is_(None),
    ).all()
    for c in charge_rows:
        item = InvoiceItemCreate(description=c.head, quantity=1, unit_price=float(c.amount), charge_id=c.id)
        if c.is_mess_charge:
            mess_items.append(item)
            amounts["other_mess_charges"] += float(c.amount)
        else:
            room_items.append(item)
            amounts["room_charges"] += float(c.amount)

    return {
        "room_items": room_items, "mess_items": mess_items, "unpriced": unpriced,
        "attendance_rows": attendance_rows, "ala_carte_orders": ala_carte_orders,
        "charge_rows": charge_rows, "amounts": amounts, "booking": booking,
    }


def _guest_display_name(booking: Booking) -> str:
    """Rank-prefixed guest name, without doubling the rank when the name was
    entered already carrying it ('Brig Nasir Iqbal' + rank 'Brig')."""
    name = booking.guest_name or "-"
    rank = (booking.rank or "").strip()
    if rank and not name.lower().startswith(rank.lower()):
        return f"{rank} {name}"
    return name


def _booking_bill_header(booking: Booking) -> dict:
    """Everything the printable draft-bill header needs, straight from the
    paper format: Online V/No, PA No, Rank, Name, Room No, Address."""
    return {
        "guest_name": booking.guest_name, "rank": booking.rank,
        "pa_number": booking.pa_number, "unit_address": booking.unit_address,
        "room_number": booking.room.room_number if booking.room else None,
        "check_in": booking.check_in, "check_out": booking.check_out,
        "reference_person": booking.reference_person,
        "source": booking.source or "walk_in", "online_voucher_no": booking.online_voucher_no,
        "advance_payment_amount": float(booking.advance_payment_amount) if booking.advance_payment_amount else 0,
        "booking_reference": booking.booking_reference,
    }


def _running_balance_payload(db: Session, booking: Booking) -> dict:
    """Read-only unbilled balance for one booking - never mutates anything.
    Shared by the single-booking endpoint and the Clerk Desk worklist."""
    gathered = _gather_unbilled_items(db, booking.id)
    amounts = gathered["amounts"]
    booking_id = booking.id
    invoices = db.query(Invoice).filter(
        Invoice.booking_id == booking_id, Invoice.status != InvoiceStatus.VOID,
    ).all()
    billed_types = {inv.bill_type for inv in invoices}
    room_billed, mess_billed = "room" in billed_types, "mess" in billed_types
    # A billed side's charges are already on an invoice - drop that side from
    # the unbilled preview instead of re-showing (and re-summing) it.
    room_items = [] if room_billed else gathered["room_items"]
    mess_items = [] if mess_billed else gathered["mess_items"]
    room_bill = 0.0 if room_billed else amounts["room"] + amounts["room_charges"]
    mess_bill = 0.0 if mess_billed else amounts["mess_total"] + amounts["gas_total"] + amounts["other_mess_charges"]
    # Online bookings pay the room charge in full in advance - netted out of
    # the preview here (capped at what's actually owed) so this preview stays
    # equal to what instant-checkout will actually charge/credit; the same
    # advance gets applied for real as a payment once the room invoice exists.
    advance_credit_applied = 0.0
    if not room_billed and booking.source == "online" and booking.advance_payment_amount:
        advance_credit_applied = min(float(booking.advance_payment_amount), room_bill)
        room_bill -= advance_credit_applied
    # What the guest still owes on bills that already exist but aren't paid off.
    outstanding = sum(float(i.total_amount) - float(i.amount_paid) for i in invoices)
    # Itemized previews for the Clerk Desk's Room/Food correction dialogs -
    # component_key marks a computed rate line as individually correctable,
    # charge_id marks an ad-hoc BookingCharge as removable; a line with
    # neither (Extra Mattress, Late Checkout Fee) is fixed/structural.
    preview = lambda item: {  # noqa: E731
        "description": item.description, "amount": round(item.unit_price * item.quantity, 2),
        "component_key": item.component_key, "charge_id": item.charge_id,
    }
    # Checkout-readiness context (informational only, doesn't gate checkout) -
    # who's already confirmed this guest's charges are final on their side.
    # See Booking.kitchen_finalized_at's model docstring.
    kitchen_finalizer = db.query(User).filter(User.id == booking.kitchen_finalized_by).first() if booking.kitchen_finalized_by else None
    booking_finalizer = db.query(User).filter(User.id == booking.booking_finalized_by).first() if booking.booking_finalized_by else None
    return {
        "room_amount": amounts["room"], "mess_charge_amount": amounts["mess_total"],
        "gas_charge_amount": amounts["gas_total"],
        "room_charges_amount": amounts["room_charges"], "other_mess_charges_amount": amounts["other_mess_charges"],
        "room_bill_total": room_bill, "mess_bill_total": mess_bill,
        "room_items": [preview(i) for i in room_items],
        "mess_items": [preview(i) for i in mess_items],
        # No pre-payment concept, so what's owed = everything not yet billed
        # plus the unpaid balance of any bill already generated.
        "total": room_bill + mess_bill,
        "outstanding_invoices": outstanding,
        "balance_due": room_bill + mess_bill + outstanding,
        "unpriced_items": gathered["unpriced"],
        "room_billed": room_billed, "mess_billed": mess_billed,
        "advance_credit_applied": advance_credit_applied,
        "kitchen_finalized_at": booking.kitchen_finalized_at, "kitchen_finalized_by_name": kitchen_finalizer.full_name if kitchen_finalizer else None,
        "booking_finalized_at": booking.booking_finalized_at, "booking_finalized_by_name": booking_finalizer.full_name if booking_finalizer else None,
    }


@router.get("/bookings/{booking_id}/running-balance")
async def running_balance(booking_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    booking = db.query(Booking).filter(Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    return _running_balance_payload(db, booking)


# --- Walk-in mess-only guests (meals consumed, no room booking) ---

def _gather_walkin_items(db: Session, guest: Guest, *, mess_override: float = None, gas_override: float = None):
    """Everything not yet invoiced for a standalone walk-in guest, consolidated
    the same way as _gather_unbilled_items: one "Extra Messing" line (a
    walk-in has no ala-carte - KitchenOrder carries no guest_id, only routine
    MealAttendance) and one "Sui Gas Charges on Messing" line."""
    mess_items, unpriced = [], []
    meal_multiplier = _walkin_meal_multiplier(db)
    mess_total, mess_unpriced, attendance_rows, _ = compute_unbilled_mess_total(
        db, guest_id=guest.id, meal_multiplier=meal_multiplier,
    )
    unpriced.extend(mess_unpriced)
    mess_total = mess_override if mess_override is not None else mess_total
    if mess_total > 0:
        mess_items.append(InvoiceItemCreate(description="Extra Messing", quantity=1, unit_price=mess_total))

    gas_total = gas_override if gas_override is not None else compute_unbilled_gas_total(
        db, attendance_rows, meal_multiplier=meal_multiplier, fallback_percentage=_gas_percentage(db),
    )
    if gas_total > 0:
        mess_items.append(InvoiceItemCreate(description="Sui Gas Charges on Messing", quantity=1, unit_price=gas_total))

    return {"mess_items": mess_items, "unpriced": unpriced, "attendance_rows": attendance_rows,
            "mess_total": mess_total, "gas_total": gas_total}


def _guest_bill_header(guest: Guest) -> dict:
    """Printable-bill header for a walk-in guest - name/address only, none of
    the room/PA/voucher fields a room stay carries (they don't exist here)."""
    return {
        "guest_name": guest.full_name, "rank": None,
        "pa_number": None, "unit_address": guest.unit_address,
        "room_number": None, "check_in": None, "check_out": None,
        "reference_person": None, "source": "walk_in", "online_voucher_no": None,
        "advance_payment_amount": 0,
        "booking_reference": f"WALKIN-{guest.id}",
    }


def _guest_running_balance_payload(db: Session, guest: Guest) -> dict:
    """Read-only unbilled mess balance for one walk-in guest. Mirrors the
    booking running-balance shape (room side always empty/zero) so the same
    Clerk Desk card/checkout components render it unchanged."""
    gathered = _gather_walkin_items(db, guest)
    invoices = db.query(Invoice).filter(
        Invoice.guest_id == guest.id, Invoice.status != InvoiceStatus.VOID,
    ).all()
    mess_billed = any(inv.bill_type == "mess" for inv in invoices)
    mess_items = [] if mess_billed else gathered["mess_items"]
    mess_bill = 0.0 if mess_billed else gathered["mess_total"] + gathered["gas_total"]
    outstanding = sum(float(i.total_amount) - float(i.amount_paid) for i in invoices)
    preview = lambda item: {  # noqa: E731
        "description": item.description,
        "amount": round(item.unit_price * item.quantity, 2),
    }
    return {
        "room_amount": 0.0, "mess_charge_amount": gathered["mess_total"],
        "gas_charge_amount": gathered["gas_total"], "room_charges_amount": 0.0, "other_mess_charges_amount": 0.0,
        "room_bill_total": 0.0, "mess_bill_total": mess_bill,
        "room_items": [], "mess_items": [preview(i) for i in mess_items],
        "total": mess_bill, "outstanding_invoices": outstanding,
        "balance_due": mess_bill + outstanding,
        "unpriced_items": gathered["unpriced"],
        "room_billed": True, "mess_billed": mess_billed,
    }


@router.get("/guests/{guest_id}/running-balance")
async def guest_running_balance(guest_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    guest = db.query(Guest).filter(Guest.id == guest_id).first()
    if not guest:
        raise HTTPException(status_code=404, detail="Guest not found")
    return _guest_running_balance_payload(db, guest)


@router.post("/guests/{guest_id}/mess-bill")
async def generate_guest_mess_bill(
    guest_id: int, request: Request,
    mess_charge_override: Optional[float] = Body(default=None, embed=True),
    gas_charge_override: Optional[float] = Body(default=None, embed=True),
    db: Session = Depends(get_db), current_user=Depends(get_current_user),
):
    """Compiles a walk-in guest's unbilled meals into one mess invoice - the
    Mess-Only page's equivalent of Instant Checkout. Clerk-only (generating a
    bill is the Clerk's job); Booking/Kitchen staff only log the meals.
    mess_charge_override/gas_charge_override replace the computed Extra
    Messing/Sui Gas totals when the clerk adjusts them in the checkout preview."""
    if not check_permission(current_user, "clerk_desk", "create"):
        raise HTTPException(status_code=403, detail="Permission denied")
    guest = db.query(Guest).filter(Guest.id == guest_id).first()
    if not guest:
        raise HTTPException(status_code=404, detail="Guest not found")

    gathered = _gather_walkin_items(db, guest, mess_override=mess_charge_override, gas_override=gas_charge_override)
    if not gathered["mess_items"]:
        raise HTTPException(status_code=400, detail="No unbilled meals to invoice for this guest")

    today = date.today()
    invoice = _build_invoice(db, None, gathered["mess_items"], current_user, today, today, bill_type="mess", guest=guest)
    now = datetime.utcnow()
    for a in gathered["attendance_rows"]:
        a.invoiced_at = now
    db.commit()
    log_audit(db, current_user.id, current_user.full_name, AuditAction.CREATE, "invoices", invoice.id,
              after_state=serialize_model(invoice), ip_address=request.client.host)
    return {
        "invoices": [_invoice_out(invoice)],
        "guest": _guest_bill_header(guest),
        "unpriced_items": gathered["unpriced"],
        "grand_total": float(invoice.total_amount),
        "balance_due": float(invoice.total_amount),
    }


@router.get("/mess-only-desk")
async def mess_only_desk(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """The Mess-Only worklist: every walk-in guest with unbilled meals (each
    with their running mess balance), plus every unsettled guest-only mess
    invoice. Guest-only invoices are kept OFF the main /desk feed and live
    here instead, so the two pages don't double-list the same bill."""
    if not (check_permission(current_user, "clerk_desk", "view") or check_permission(current_user, "billing", "view")):
        raise HTTPException(status_code=403, detail="Permission denied")
    today = date.today()
    # Guests with at least one unbilled, billable meal and no active room stay
    # (a checked-in guest bills through checkout, not here).
    checked_in_guest_ids = {gid for (gid,) in db.query(Booking.guest_id).filter(
        Booking.status == "checked_in", Booking.guest_id.isnot(None)).all()}
    guest_ids = {gid for (gid,) in db.query(MealAttendance.guest_id).filter(
        MealAttendance.guest_id.isnot(None), MealAttendance.invoiced_at.is_(None),
        MealAttendance.status.in_(["booked", "attended", "no_show"]),
    ).all()} - checked_in_guest_ids
    guests = db.query(Guest).filter(Guest.id.in_(guest_ids)).all() if guest_ids else []

    items = []
    for g in guests:
        bal = _guest_running_balance_payload(db, g)
        if bal["mess_bill_total"] <= 0.01 and not bal["unpriced_items"]:
            continue  # nothing priced to bill
        items.append({
            "id": g.id, "guest_name": g.full_name, "rank": None,
            "room_number": None, "status": "walk_in", "balance": bal,
        })

    # Event invoices are also guest_id-linked (no booking) but aren't a walk-in
    # mess bill - they'd otherwise show up here mislabeled as "mess". They're
    # settled from the Events page or the general Billing page instead.
    open_invoices = db.query(Invoice).filter(
        Invoice.guest_id.isnot(None), Invoice.status.in_((InvoiceStatus.DRAFT, InvoiceStatus.ISSUED)),
        Invoice.bill_type != "event",
    ).order_by(Invoice.created_at.desc()).all()
    unsettled = []
    for inv in open_invoices:
        balance = float(inv.total_amount) - float(inv.amount_paid)
        if balance <= 0.01:
            continue
        g = inv.guest
        unsettled.append({
            "id": inv.id, "invoice_number": inv.invoice_number, "bill_type": "mess",
            "total_amount": float(inv.total_amount), "amount_paid": float(inv.amount_paid),
            "balance_due": balance, "guest_id": inv.guest_id,
            "guest_name": g.full_name if g else None, "rank": None, "room_number": None,
            "issue_date": inv.issue_date, "checking_out_now": False,
            "overdue": bool(inv.due_date and inv.due_date < today and balance > 0.01),
            "issued_today": bool(inv.issue_date == today),
        })
    return {"items": items, "unsettled_invoices": unsettled}


@router.get("/desk")
async def clerk_desk(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """The Clerk Desk worklist in one request: every guest who still needs
    billing attention, each with their running balance embedded.

    - checked-in guests (HRA residents excluded - they settle via the
      monthly Mess Bill and must never be checkout-able here), and
    - every generated-but-unsettled invoice, guests checking out TODAY
      first - so payment collection lives entirely on this desk instead of
      requiring a trip to the Billing page when the print dialog was
      closed without recording payment.

    Checkout always bills everything owed in one shot (walk-ins pay room+mess
    together on the spot; online guests prepay the room in advance and pay
    mess at checkout) - there's no "checked out but still unbilled" worklist
    here by design. A booking only reappears after checkout via
    unsettled_invoices, if its bill was generated but not fully paid."""
    if not (check_permission(current_user, "clerk_desk", "view") or check_permission(current_user, "billing", "view")):
        raise HTTPException(status_code=403, detail="Permission denied")
    checked_in = db.query(Booking).options(joinedload(Booking.room)).filter(
        Booking.status == "checked_in",
        or_(Booking.nature_of_duty.is_(None), Booking.nature_of_duty != "hra"),
    ).order_by(Booking.created_at.desc()).all()

    # Unsettled bills: anything live with money still owing. Legacy invoices
    # predate the created-as-issued rule, so 'draft' is included too. Guest-only
    # walk-in mess invoices are excluded here - they live on the Mess-Only page.
    today = date.today()
    open_invoices = db.query(Invoice).options(joinedload(Invoice.booking).joinedload(Booking.room)).filter(
        Invoice.status.in_((InvoiceStatus.DRAFT, InvoiceStatus.ISSUED)),
        Invoice.guest_id.is_(None),
    ).order_by(Invoice.created_at.desc()).all()
    unsettled = []
    for inv in open_invoices:
        balance = float(inv.total_amount) - float(inv.amount_paid)
        if balance <= 0.01:
            continue
        b = inv.booking
        unsettled.append({
            "id": inv.id, "invoice_number": inv.invoice_number,
            "bill_type": inv.bill_type or "combined",
            "total_amount": float(inv.total_amount), "amount_paid": float(inv.amount_paid),
            "balance_due": balance,
            "booking_id": inv.booking_id,
            "guest_name": b.guest_name if b else None, "rank": b.rank if b else None,
            "room_number": b.room.room_number if b and b.room else None,
            "issue_date": inv.issue_date,
            # The guest is standing at the desk right now - settle these first.
            "checking_out_now": bool(b and b.actual_check_out and b.actual_check_out.date() == today),
            # State flags so the frontend can colour without re-deriving.
            "overdue": bool(inv.due_date and inv.due_date < today and balance > 0.01),
            "issued_today": bool(inv.issue_date == today),
        })
    unsettled.sort(key=lambda r: not r["checking_out_now"])  # stable: keeps newest-first within each group

    return {
        "items": [{
            "id": b.id, "booking_reference": b.booking_reference,
            "guest_name": b.guest_name, "rank": b.rank,
            "room_number": b.room.room_number if b.room else None,
            "status": b.status.value, "source": b.source or "walk_in",
            "balance": _running_balance_payload(db, b),
        } for b in checked_in],
        "unsettled_invoices": unsettled,
    }


# --- Ad-hoc booking charges (Dhobi, Breakage, Allied, Extra Messing...) ---

@router.get("/bookings/{booking_id}/charges")
async def list_booking_charges(booking_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    # Clerk (billing:view), Booking NCO (bookings:view), or Kitchen NCO
    # (kitchen:view) - see add_booking_charge below for why each NCO is
    # scoped in on their respective side.
    if not (check_permission(current_user, "billing", "view") or check_permission(current_user, "bookings", "view")
            or check_permission(current_user, "kitchen", "view")):
        raise HTTPException(status_code=403, detail="Permission denied")
    charges = db.query(BookingCharge).filter(BookingCharge.booking_id == booking_id).order_by(BookingCharge.created_at.desc()).all()
    return [{"id": c.id, "head": c.head, "amount": float(c.amount), "is_mess_charge": c.is_mess_charge,
             "reason": c.reason, "invoiced": c.invoiced_at is not None, "created_at": c.created_at} for c in charges]


@router.post("/bookings/{booking_id}/charges")
async def add_booking_charge(booking_id: int, data: BookingChargeCreate, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    # Mess-side ad-hoc charges (a genuinely one-off Extra Messing/Sui Gas
    # entry, not the normal auto-computed flow) are owned end-to-end by
    # whichever of Clerk (mess_billing:create) or Kitchen NCO (kitchen:edit)
    # is logging it - same "same level of access" pattern as the room side.
    # Room-side charges (Dhobi/Breakage/Wages of Servants/Heater-AC/etc.) are
    # owned end-to-end by whichever of Clerk (billing:create) or Booking NCO
    # (bookings:edit) is logging it - both NCOs already own their respective
    # module for the underlying stay/kitchen operation, so no separate
    # billing/mess_billing grant is needed just to log its ad-hoc charges.
    if data.is_mess_charge:
        if not (check_permission(current_user, "mess_billing", "create") or check_permission(current_user, "kitchen", "edit")):
            raise HTTPException(status_code=403, detail="Permission denied")
    else:
        if not (check_permission(current_user, "billing", "create") or check_permission(current_user, "bookings", "edit")):
            raise HTTPException(status_code=403, detail="Permission denied")
    booking = db.query(Booking).filter(Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking.status.value not in ("confirmed", "checked_in"):
        raise HTTPException(status_code=400, detail=f"Cannot add charges to a booking with status '{booking.status.value}'")
    # Reason is only mandatory for the ambiguous "Allied Charges" catch-all -
    # every other preset head is already self-explanatory from its name.
    if data.head.strip() == "Allied Charges" and len((data.reason or "").strip()) < 3:
        raise HTTPException(status_code=400, detail="A reason (at least 3 characters) is required for Allied Charges")
    charge = BookingCharge(booking_id=booking_id, head=data.head.strip(), amount=data.amount,
                           is_mess_charge=data.is_mess_charge, reason=(data.reason or "").strip() or None, created_by=current_user.id)
    db.add(charge)
    db.commit()
    db.refresh(charge)
    log_audit(db, current_user.id, current_user.full_name, AuditAction.CREATE, "booking_charges", charge.id,
              after_state=serialize_model(charge), reason=charge.reason, ip_address=request.client.host)
    return {"id": charge.id, "head": charge.head, "amount": float(charge.amount), "is_mess_charge": charge.is_mess_charge, "reason": charge.reason}


@router.delete("/charges/{charge_id}")
async def delete_booking_charge(charge_id: int, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    charge = db.query(BookingCharge).filter(BookingCharge.id == charge_id).first()
    if not charge:
        raise HTTPException(status_code=404, detail="Charge not found")
    if charge.is_mess_charge:
        if not (check_permission(current_user, "mess_billing", "edit") or check_permission(current_user, "kitchen", "edit")):
            raise HTTPException(status_code=403, detail="Permission denied")
    else:
        if not (check_permission(current_user, "billing", "edit") or check_permission(current_user, "bookings", "edit")):
            raise HTTPException(status_code=403, detail="Permission denied")
    if charge.invoiced_at is not None:
        raise HTTPException(status_code=400, detail="Charge is already on an invoice - void the invoice instead")
    before = serialize_model(charge)
    db.delete(charge)
    db.commit()
    log_audit(db, current_user.id, current_user.full_name, AuditAction.SOFT_DELETE, "booking_charges", charge_id,
              before_state=before, reason="Charge removed before invoicing", ip_address=request.client.host)
    return {"message": "Charge removed"}


def _invoice_out(inv: Invoice) -> dict:
    return {
        "id": inv.id, "invoice_number": inv.invoice_number, "bill_type": inv.bill_type or "combined",
        "issue_date": inv.issue_date, "subtotal": float(inv.subtotal),
        "total_amount": float(inv.total_amount), "amount_paid": float(inv.amount_paid),
        "balance_due": float(inv.total_amount) - float(inv.amount_paid),
        "status": inv.status.value, "is_complimentary": bool(inv.is_complimentary),
        "bill_serial_number": inv.bill_serial_number,
        "items": [{"id": i.id, "description": i.description, "quantity": float(i.quantity),
                   "unit_price": float(i.unit_price), "total_price": float(i.total_price)} for i in inv.items],
    }


@router.post("/bookings/{booking_id}/instant-checkout")
async def instant_checkout(
    booking_id: int, request: Request,
    bill_types: Optional[List[str]] = Body(default=None, embed=True),
    mess_charge_override: Optional[float] = Body(default=None, embed=True),
    gas_charge_override: Optional[float] = Body(default=None, embed=True),
    room_component_overrides: Optional[Dict[str, float]] = Body(default=None, embed=True),
    room_override_reason: Optional[str] = Body(default=None, embed=True),
    db: Session = Depends(get_db), current_user=Depends(get_current_user),
):
    """Compiles all unbilled charges for a checked-in guest - room charges
    (itemized rate components - rent/electricity/etc, see
    ROOM_COMPONENT_LABELS - + dhobi/breakage/allied...) and mess charges (one
    consolidated Extra Messing line + one Sui Gas Charges line, computed from
    everything ordered - see mess_charge_calc.compute_unbilled_mess_total)
    into a room invoice and/or a mess invoice. bill_types selects which side
    to include now - defaults to both, but either can be settled alone and
    the other finalized later (e.g. an a la carte order still pending).
    mess_charge_override/gas_charge_override replace the computed totals when
    the clerk adjusts them in the checkout preview. room_component_overrides
    replaces one or more individual rate-component lines (keyed the same as
    compute_booking_price's components dict, e.g. "rent"/"electricity") -
    Extra Mattress/Late Checkout Fee/any ad-hoc BookingCharge stay separate
    lines regardless, so the bill never collapses into one flat number - and
    requires room_override_reason (non-empty) to be supplied alongside it.
    When both sides are billed in the same call they land on two separate
    invoices (kept detailed - room-only and mess-only rows, not merged), and
    the Clerk's print dialog offers a Room / Mess / Combined toggle over them
    (the combined view row-merges both via GET .../master-invoice). No
    advance is applied - payment is settled here at checkout, via the Pay
    Together / per-bill payment actions the print dialog offers. Guest/
    booking-only - members settle through the monthly Mess Bill cycle instead."""
    booking = db.query(Booking).filter(Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    if booking.status.value not in ("checked_in", "checked_out"):
        raise HTTPException(status_code=400, detail=f"Cannot check out a booking with status '{booking.status.value}' - guest must be checked in")
    if booking.nature_of_duty == "hra":
        raise HTTPException(status_code=400, detail="HRA residents settle via the monthly Mess Bill - this residency cannot be checked out here")
    if room_component_overrides and not (room_override_reason or "").strip():
        raise HTTPException(status_code=400, detail="A reason is required to override a room charge component")

    # billing/mess_billing:create only ever logs a charge (add_booking_charge,
    # gated separately above) - generating the actual invoice needs its own,
    # separate clerk_desk:create permission.
    if not check_permission(current_user, "clerk_desk", "create"):
        raise HTTPException(status_code=403, detail="Permission denied")

    requested = set(bill_types) if bill_types else {"room", "mess"}
    if requested - {"room", "mess"}:
        raise HTTPException(status_code=400, detail="bill_types must be 'room' and/or 'mess'")
    # 'combined' (this endpoint's normal output now) covers both sides at
    # once; legacy 'room'/'mess' rows (from before bills were merged) each
    # cover only their own side.
    already_billed_types = {inv.bill_type for inv in db.query(Invoice).filter(
        Invoice.booking_id == booking_id, Invoice.status != InvoiceStatus.VOID,
    ).all()}
    room_done = bool(already_billed_types & {"room", "combined"})
    mess_done = bool(already_billed_types & {"mess", "combined"})
    # Silently skip a side that's already billed instead of failing the whole
    # request - Clerk always asks for "both", and one side may have been
    # settled earlier (e.g. an a la carte order still pending at the time).
    if room_done:
        requested.discard("room")
    if mess_done:
        requested.discard("mess")

    # Actually check the guest out FIRST (status, room freed to housekeeping,
    # late fee assessed) so the fee lands on the room bill and the Clerk Desk
    # card disappears. Previously this endpoint only invoiced, leaving the
    # booking checked_in - the card stayed put and a second click 409'd.
    late_fee = 0.0
    if booking.status.value == "checked_in":
        from backend.routers.bookings import perform_check_out
        late_fee = perform_check_out(db, booking, current_user)

    gathered = _gather_unbilled_items(db, booking_id, mess_override=mess_charge_override, gas_override=gas_charge_override, room_component_overrides=room_component_overrides)
    room_items = gathered["room_items"] if "room" in requested else []
    mess_items = gathered["mess_items"] if "mess" in requested else []
    if not room_items and not mess_items:
        raise HTTPException(status_code=400, detail="Nothing to invoice for this booking")

    today = date.today()
    now = datetime.utcnow()
    # Room and mess stay on separate invoices even when billed in the same
    # call, so each keeps its own detailed rows (Guest Room Charges / Extra
    # Mattress on one side, Extra Messing / Sui Gas / per-meal rows on the
    # other) - the print dialog's Combined view row-merges them for display
    # without losing that per-invoice detail.
    invoices = []
    if room_items:
        room_notes = None
        if room_component_overrides:
            overridden = ", ".join(
                f"{ROOM_COMPONENT_LABELS.get(k, k.replace('_', ' ').title())} to Rs {v:,.0f}"
                for k, v in room_component_overrides.items()
            )
            room_notes = f"{overridden} - {room_override_reason.strip()}"
        room_invoice = _build_invoice(db, booking, room_items, current_user, today, today, notes=room_notes, bill_type="room")
        invoices.append(room_invoice)
        for c in gathered["charge_rows"]:
            if not c.is_mess_charge:
                c.invoiced_at = now
        # Online bookings paid the room charge in full, in advance, outside
        # SAM - credit it now that the room invoice exists, dated to when it
        # was ACTUALLY received (not today), so Today's/Month's Collections
        # attribute it to the right day instead of looking like cash just
        # came in at checkout. Capped at the invoice total defensively, even
        # though online advances are always sized to match the room charge.
        if booking.source == "online" and booking.advance_payment_amount:
            gross_advance = float(booking.advance_payment_amount)
            advance = min(gross_advance, float(room_invoice.total_amount))
            if advance > 0:
                received_at = datetime.combine(booking.advance_paid_at, datetime.min.time()) if booking.advance_paid_at else now
                # AG Branch fee is 10% of the GROSS advance (what the guest
                # actually paid online), independent of `advance` above -
                # the guest is still credited the full (capped) amount; this
                # is purely what AG Branch retains from the mess's own net
                # payout, tracked for reporting only.
                ag_fee = gross_advance * get_setting_float(db, "ag_branch_fee_percentage", 10) / 100
                db.add(InvoicePayment(
                    invoice_id=room_invoice.id, amount=advance, method=PAYMENT_METHOD_ONLINE,
                    notes=f"Online booking advance - V/No {booking.online_voucher_no or '—'}",
                    voucher_number=booking.online_voucher_no, ag_branch_fee=ag_fee,
                    received_by=current_user.id, created_at=received_at,
                ))
                room_invoice.amount_paid = advance
                if advance >= float(room_invoice.total_amount) + float(room_invoice.last_debit_balance or 0) - 0.01:
                    room_invoice.status = InvoiceStatus.PAID
                elif advance > 0:
                    room_invoice.status = InvoiceStatus.PARTIALLY_PAID
    if mess_items:
        invoices.append(_build_invoice(db, booking, mess_items, current_user, today, today, bill_type="mess"))
        for a in gathered["attendance_rows"]:
            a.invoiced_at = now
        for o in gathered["ala_carte_orders"]:
            o.invoiced_at = now
        for c in gathered["charge_rows"]:
            if c.is_mess_charge:
                c.invoiced_at = now
    db.commit()

    for inv in invoices:
        inv_reason = room_override_reason if (inv.bill_type == "room" and room_component_overrides) else None
        log_audit(db, current_user.id, current_user.full_name, AuditAction.CREATE, "invoices", inv.id,
                  after_state=serialize_model(inv), reason=inv_reason, ip_address=request.client.host)

    grand_total = sum(float(i.total_amount) for i in invoices)
    return {
        "invoices": [_invoice_out(inv) for inv in invoices],
        "booking": _booking_bill_header(booking),
        "unpriced_items": gathered["unpriced"],
        "late_checkout_fee": late_fee,
        "grand_total": grand_total,
        "balance_due": grand_total,
    }


def _invoice_qr_svg(payload: str) -> str:
    """Locally generated QR (reportlab - no internet, nothing leaves the
    LAN), returned as an inline-able SVG string."""
    from reportlab.graphics.barcode.qr import QrCodeWidget
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderSVG
    qr = QrCodeWidget(payload)
    x0, y0, x1, y1 = qr.getBounds()
    size = 200.0
    drawing = Drawing(size, size, transform=[size / (x1 - x0), 0, 0, size / (y1 - y0), 0, 0])
    drawing.add(qr)
    svg = renderSVG.drawToString(drawing)
    return svg[svg.find("<svg"):]  # strip the XML declaration so it can be inlined in HTML


@router.get("/bookings/{booking_id}/master-invoice")
async def master_invoice(booking_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Clerk consolidation: combines a stay's separate room and mess
    invoices (generated independently by Booking Staff and Mess Staff) into
    one comprehensive master invoice - one merged item list, one grand
    total, one printable document. Read-only; the underlying invoices stay
    exactly as they are, so nothing here can double-count revenue."""
    if not check_permission(current_user, "clerk_desk", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    booking = db.query(Booking).filter(Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    invoices = db.query(Invoice).filter(
        Invoice.booking_id == booking_id, Invoice.status != InvoiceStatus.VOID,
    ).order_by(Invoice.issue_date, Invoice.id).all()
    if not invoices:
        raise HTTPException(status_code=404, detail="No bills generated yet for this stay")

    items = []
    for inv in invoices:
        label = {"room": "Room Bill", "mess": "Mess Bill"}.get(inv.bill_type or "combined", "Bill")
        for it in inv.items:
            items.append({
                "source": inv.bill_type or "combined", "source_label": label,
                "description": it.description, "quantity": float(it.quantity),
                "unit_price": float(it.unit_price), "total_price": float(it.total_price),
            })

    subtotal = sum(float(inv.subtotal) for inv in invoices)
    tax_amount = sum(float(inv.tax_amount) for inv in invoices)
    discount = sum(float(inv.discount) for inv in invoices)
    total_amount = sum(float(inv.total_amount) for inv in invoices)
    amount_paid = sum(float(inv.amount_paid) for inv in invoices)

    mess_name = get_setting_str(db, "mess_name", "EME Officers Mess")
    mess_address = get_setting_str(db, "mess_address", "204 Firdousi Road, Rawalpindi")
    mess_phone = get_setting_str(db, "mess_phone", "Tele No. G.H.Q 31725")
    verify_hash = hashlib.sha256(
        "|".join([str(inv.id) for inv in invoices] + [f"{total_amount:.2f}"]).encode()
    ).hexdigest()[:12].upper()

    return {
        "booking": _booking_bill_header(booking),
        "source_invoices": [{"id": inv.id, "invoice_number": inv.invoice_number, "bill_type": inv.bill_type or "combined"} for inv in invoices],
        "items": items,
        "subtotal": subtotal, "tax_amount": tax_amount, "discount": discount,
        "total_amount": total_amount, "amount_paid": amount_paid, "balance_due": total_amount - amount_paid,
        "is_complimentary": all(inv.is_complimentary for inv in invoices),
        "mess": {"name": mess_name, "address": mess_address, "phone": mess_phone},
        "verify_hash": verify_hash,
        "qr_svg": _invoice_qr_svg(f"{mess_name}\nMaster Invoice: {', '.join(i.invoice_number for i in invoices)}\nTotal: Rs {total_amount:,.0f}\nVerify: {verify_hash}"),
    }


def _invoice_edit_request_out(db: Session, req: InvoiceEditRequest) -> dict:
    inv = req.invoice
    booking = inv.booking if inv else None
    requester = db.query(User).filter(User.id == req.requested_by).first() if req.requested_by else None
    decider = db.query(User).filter(User.id == req.decided_by).first() if req.decided_by else None
    return {
        "id": req.id, "invoice_id": req.invoice_id, "invoice_item_id": req.invoice_item_id,
        "bill_type": inv.bill_type if inv else "combined",
        "original_description": req.original_description, "original_unit_price": float(req.original_unit_price),
        "proposed_description": req.proposed_description, "proposed_unit_price": float(req.proposed_unit_price),
        "reason": req.reason, "status": req.status.value,
        "requested_by_name": requester.full_name if requester else None,
        "requested_at": req.requested_at,
        "decided_by_name": decider.full_name if decider else None,
        "decided_at": req.decided_at, "decision_reason": req.decision_reason,
        "guest_name": booking.guest_name if booking else None,
        "room_number": booking.room.room_number if booking and booking.room else None,
    }


@router.post("/invoice-items/{item_id}/edit-request")
async def request_invoice_item_edit(item_id: int, data: InvoiceEditRequestCreate, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """A Clerk-proposed correction to one line on an already-generated bill
    (wrong room rate, wrong mess charge...) - sits pending until a Manager
    approves (POST .../approve) or rejects (POST .../reject) it. Scoped to
    description/unit_price only; quantity is untouched. Distinct from the
    discount/complimentary action on Invoice, which stays Clerk-autonomous
    with no approval step."""
    if not check_permission(current_user, "billing", "edit"):
        raise HTTPException(status_code=403, detail="Permission denied")
    item = db.query(InvoiceItem).filter(InvoiceItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Invoice item not found")
    inv = item.invoice
    if inv.status in (InvoiceStatus.VOID, InvoiceStatus.PAID):
        raise HTTPException(status_code=400, detail=f"Cannot request a correction on an invoice with status '{inv.status.value}'")
    existing = db.query(InvoiceEditRequest).filter(
        InvoiceEditRequest.invoice_item_id == item_id, InvoiceEditRequest.status == EditRequestStatus.PENDING,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="This line already has a correction pending approval")

    req = InvoiceEditRequest(
        invoice_item_id=item.id, invoice_id=inv.id,
        original_description=item.description, original_unit_price=item.unit_price,
        proposed_description=data.proposed_description, proposed_unit_price=data.proposed_unit_price,
        reason=data.reason, requested_by=current_user.id,
    )
    db.add(req)
    db.commit()
    db.refresh(req)
    log_audit(db, current_user.id, current_user.full_name, AuditAction.CREATE, "invoice_edit_requests", req.id,
              after_state=serialize_model(req), ip_address=request.client.host)
    return _invoice_edit_request_out(db, req)


@router.post("/invoices/{invoice_id}/edit-request")
async def request_invoice_new_line(invoice_id: int, data: InvoiceEditRequestCreate, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Same correction workflow as request_invoice_item_edit above, but for
    a head that's currently zero/uncharged rather than an existing line -
    e.g. the paper form's "Dhobi" row was never billed. invoice_item_id
    stays null; on approval a new InvoiceItem is created instead of an
    existing one being updated."""
    if not check_permission(current_user, "billing", "edit"):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.status in (InvoiceStatus.VOID, InvoiceStatus.PAID):
        raise HTTPException(status_code=400, detail=f"Cannot request a correction on an invoice with status '{inv.status.value}'")

    req = InvoiceEditRequest(
        invoice_item_id=None, invoice_id=inv.id,
        original_description="Not yet charged", original_unit_price=0,
        proposed_description=data.proposed_description, proposed_unit_price=data.proposed_unit_price,
        reason=data.reason, requested_by=current_user.id,
    )
    db.add(req)
    db.commit()
    db.refresh(req)
    log_audit(db, current_user.id, current_user.full_name, AuditAction.CREATE, "invoice_edit_requests", req.id,
              after_state=serialize_model(req), ip_address=request.client.host)
    return _invoice_edit_request_out(db, req)


@router.get("/edit-requests")
async def list_edit_requests(status: str = "pending", db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Used both by the Clerk (their own submitted requests) and the Manager
    (the approval queue) - whichever side of the workflow the caller has
    permission for."""
    if not (check_permission(current_user, "billing", "edit") or check_permission(current_user, "billing", "approve")):
        raise HTTPException(status_code=403, detail="Permission denied")
    query = db.query(InvoiceEditRequest)
    if status:
        try:
            query = query.filter(InvoiceEditRequest.status == EditRequestStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail="status must be pending, approved, or rejected")
    reqs = query.order_by(InvoiceEditRequest.requested_at.desc()).all()
    return [_invoice_edit_request_out(db, r) for r in reqs]


def _recompute_invoice_totals(inv: Invoice):
    subtotal = sum(float(i.total_price) for i in inv.items)
    inv.subtotal = subtotal
    inv.total_amount = subtotal + float(inv.tax_amount) - float(inv.discount)


@router.post("/edit-requests/{request_id}/approve")
async def approve_edit_request(request_id: int, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Manager-only: applies the proposed correction and recomputes the
    parent invoice's subtotal/total from all its items. A request with no
    invoice_item_id (added under a zero/uncharged head) creates a new
    InvoiceItem instead of updating an existing one."""
    if not check_permission(current_user, "billing", "approve"):
        raise HTTPException(status_code=403, detail="Permission denied")
    req = db.query(InvoiceEditRequest).filter(InvoiceEditRequest.id == request_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Edit request not found")
    if req.status != EditRequestStatus.PENDING:
        raise HTTPException(status_code=400, detail=f"Request already {req.status.value}")

    inv = req.invoice
    before_inv = serialize_model(inv)

    if req.invoice_item_id is None:
        item = InvoiceItem(
            invoice_id=inv.id, description=req.proposed_description,
            quantity=1, unit_price=float(req.proposed_unit_price), total_price=float(req.proposed_unit_price),
        )
        db.add(item)
        before_item = None
        item_audit_action = AuditAction.CREATE
    else:
        item = req.invoice_item
        before_item = serialize_model(item)
        item.description = req.proposed_description
        item.unit_price = float(req.proposed_unit_price)
        item.total_price = float(req.proposed_unit_price) * item.quantity
        item_audit_action = AuditAction.OVERRIDE
    db.flush()
    _recompute_invoice_totals(inv)

    if float(inv.total_amount) < float(inv.amount_paid) - 0.01:
        db.rollback()
        raise HTTPException(status_code=400, detail="This correction would reduce the bill below what's already been paid - void and reissue instead")

    req.status = EditRequestStatus.APPROVED
    req.decided_by = current_user.id
    req.decided_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    db.refresh(inv)
    db.refresh(req)

    log_audit(db, current_user.id, current_user.full_name, item_audit_action, "invoice_items", item.id,
              before_state=before_item, after_state=serialize_model(item), reason=req.reason, ip_address=request.client.host)
    log_audit(db, current_user.id, current_user.full_name, AuditAction.OVERRIDE, "invoices", inv.id,
              before_state=before_inv, after_state=serialize_model(inv), reason=f"Line item correction approved: {req.reason}", ip_address=request.client.host)
    return _invoice_edit_request_out(db, req)


@router.post("/edit-requests/{request_id}/reject")
async def reject_edit_request(request_id: int, data: InvoiceEditDecision, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "billing", "approve"):
        raise HTTPException(status_code=403, detail="Permission denied")
    if not data.reason:
        raise HTTPException(status_code=400, detail="A reason is required to reject a correction request")
    req = db.query(InvoiceEditRequest).filter(InvoiceEditRequest.id == request_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Edit request not found")
    if req.status != EditRequestStatus.PENDING:
        raise HTTPException(status_code=400, detail=f"Request already {req.status.value}")

    before = serialize_model(req)
    req.status = EditRequestStatus.REJECTED
    req.decided_by = current_user.id
    req.decided_at = datetime.utcnow()
    req.decision_reason = data.reason
    db.commit()
    db.refresh(req)

    log_audit(db, current_user.id, current_user.full_name, AuditAction.OVERRIDE, "invoice_edit_requests", req.id,
              before_state=before, after_state=serialize_model(req), reason=data.reason, ip_address=request.client.host)
    return _invoice_edit_request_out(db, req)


@router.get("/invoices/{invoice_id}/print-data")
async def invoice_print_data(invoice_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Everything needed to print one bill in the mess's paper format:
    invoice lines, booking-register header fields, mess identity, and a QR
    code carrying the bill summary. Guest PII and amounts, and invoice_id is
    directly enumerable - gated the same as /desk, plus guests:view since the
    Customer Directory's BillPrintView (Manager-only) also calls this."""
    if not (check_permission(current_user, "billing", "view") or check_permission(current_user, "clerk_desk", "view")
            or check_permission(current_user, "guests", "view")):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    booking = inv.booking

    mess_name = get_setting_str(db, "mess_name", "EME Officers Mess")
    mess_address = get_setting_str(db, "mess_address", "204 Firdousi Road, Rawalpindi")
    mess_phone = get_setting_str(db, "mess_phone", "Tele No. G.H.Q 31725")

    balance = float(inv.total_amount) - float(inv.amount_paid)
    bill_label = {"room": "Room Bill", "mess": "Mess Bill"}.get(inv.bill_type or "combined", "Bill")
    # Tamper-evident summary hash for audit: recomputable from the stored
    # invoice, so a printed bill can be verified against the system later.
    verify_hash = hashlib.sha256(
        f"{inv.id}|{inv.invoice_number}|{float(inv.total_amount):.2f}|{inv.issue_date.isoformat()}".encode()
    ).hexdigest()[:12].upper()
    # Header: a room stay uses the full booking-register header; a walk-in
    # mess guest uses the name-only guest header (no room/PA/voucher fields).
    header = _booking_bill_header(booking) if booking else (_guest_bill_header(inv.guest) if inv.guest else None)
    guest_line = _guest_display_name(booking) if booking else (inv.guest.full_name if inv.guest else None)
    payload = "\n".join(filter(None, [
        mess_name,
        f"{bill_label}: {inv.invoice_number}",
        "COMPLIMENTARY" if inv.is_complimentary else None,
        f"Guest: {guest_line}" if guest_line else None,
        f"Room: {booking.room.room_number}" if booking and booking.room else None,
        f"Stay: {booking.check_in.strftime('%d-%m-%y')} to {booking.check_out.strftime('%d-%m-%y')}" if booking else None,
        f"Online V/No: {booking.online_voucher_no}" if booking and booking.online_voucher_no else None,
        f"Total: Rs {float(inv.total_amount):,.0f}",
        f"Paid: Rs {float(inv.amount_paid):,.0f}",
        f"Balance: Rs {balance:,.0f}",
        f"Date: {inv.issue_date.strftime('%d-%m-%Y')}",
        f"Verify: {verify_hash}",
    ]))

    return {
        "invoice": _invoice_out(inv),
        "booking": header,
        "mess": {"name": mess_name, "address": mess_address, "phone": mess_phone},
        "verify_hash": verify_hash,
        "qr_svg": _invoice_qr_svg(payload),
    }


@router.post("/invoices/{invoice_id}/void")
async def void_invoice(invoice_id: int, reason: str, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if not check_permission(current_user, "clerk_desk", "edit"):
        raise HTTPException(status_code=403, detail="Permission denied")
    if inv.status == InvoiceStatus.VOID:
        raise HTTPException(status_code=400, detail="Invoice is already void")

    before = serialize_model(inv)
    inv.status = InvoiceStatus.VOID

    # Release the source rows this bill had claimed, so a corrected bill can
    # be generated afterwards - otherwise the meals/charges stay stamped
    # invoiced_at forever and silently vanish from any future bill.
    if inv.booking_id:
        bill_type = inv.bill_type or "combined"
        if bill_type in ("mess", "combined"):
            for a in db.query(MealAttendance).filter(
                    MealAttendance.booking_id == inv.booking_id, MealAttendance.invoiced_at.isnot(None)).all():
                a.invoiced_at = None
            for o in db.query(KitchenOrder).filter(
                    KitchenOrder.booking_id == inv.booking_id, KitchenOrder.is_ala_carte == True,
                    KitchenOrder.invoiced_at.isnot(None)).all():
                o.invoiced_at = None
        charge_query = db.query(BookingCharge).filter(
            BookingCharge.booking_id == inv.booking_id, BookingCharge.invoiced_at.isnot(None))
        if bill_type == "mess":
            charge_query = charge_query.filter(BookingCharge.is_mess_charge == True)
        elif bill_type == "room":
            charge_query = charge_query.filter(BookingCharge.is_mess_charge == False)
        for c in charge_query.all():
            c.invoiced_at = None
    db.commit()

    log_audit(db, current_user.id, current_user.full_name, AuditAction.OVERRIDE, "invoices", inv.id, before_state=before, after_state=serialize_model(inv), reason=f"Voided: {reason}", ip_address=request.client.host)
    return {"message": "Invoice voided"}


@router.put("/invoices/{invoice_id}/serial-number")
async def set_bill_serial_number(
    invoice_id: int, bill_serial_number: Optional[str] = Body(default=None, embed=True),
    request: Request = None, db: Session = Depends(get_db), current_user=Depends(get_current_user),
):
    """Clerk-entered physical bill-book serial, logged against the invoice
    before printing/settling - distinct from the system invoice_number."""
    if not (check_permission(current_user, "clerk_desk", "edit") or check_permission(current_user, "billing", "edit")):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.bill_made_at is not None:
        raise HTTPException(status_code=400, detail="This bill has been made - the serial number can no longer be changed here")
    before = serialize_model(inv)
    inv.bill_serial_number = (bill_serial_number or "").strip() or None
    db.commit()
    log_audit(db, current_user.id, current_user.full_name, AuditAction.UPDATE, "invoices", inv.id, before_state=before, after_state=serialize_model(inv), ip_address=request.client.host if request else None)
    return {"message": "Bill serial number saved", "bill_serial_number": inv.bill_serial_number}


def _invoice_master_bill_label(description: str) -> str:
    """Maps an already-persisted InvoiceItem's free-text description onto the
    paper form's fixed vocabulary (services/master_bill.py). Ad-hoc charge
    lines (Dhobi, Breakage, Masjid...) already carry their head as the
    description verbatim, so they pass through unchanged and match directly."""
    if description.startswith("Guest Room Charges"):
        return "Guest Room Charges"
    if description == "Extra Mattress":
        return "Extra Mettress"  # paper form's own spelling
    if description == "Sui Gas Charges on Messing":
        return "Sui Gas Charges on Messing / Extra Messing"
    return description


@router.get("/invoices/{invoice_id}/master-bill")
async def get_master_bill(invoice_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")

    entries = [{"label": _invoice_master_bill_label(it.description), "amount": float(it.unit_price) * it.quantity, "reason": it.reason} for it in inv.items]
    assembled = assemble_rows(entries=entries, last_debit_balance=float(inv.last_debit_balance or 0), credits=float(inv.discount or 0))

    payments = db.query(InvoicePayment).filter(InvoicePayment.invoice_id == inv.id).order_by(InvoicePayment.created_at).all()
    payment_breakdown = _build_payment_breakdown(payments)
    payment_breakdown["balance_due"] = round(assembled["net_debits"] - payment_breakdown["total_paid"], 2)

    booking = inv.booking
    header = {
        "bill_for": None,  # guest bills aren't monthly - the paper form's "Bill for the Month of" is member-only
        "rank": booking.rank if booking else None,
        "name": (booking.guest_name if booking else (inv.guest.full_name if inv.guest else None)),
        "address_unit": booking.unit_address if booking else None,
        "date": inv.issue_date,
    }
    mess = {
        "name": get_setting_str(db, "mess_name", "EME Officers Mess"),
        "address": get_setting_str(db, "mess_address", "204 Firdousi Road, Rawalpindi"),
        "phone": get_setting_str(db, "mess_phone", "Tele No. G.H.Q 31725"),
    }
    # Raw, uncollapsed items for the pre-lock Interactive correction table -
    # `rows` above is aggregated by label (several Dhobi charges become one
    # printed line) so it can't identify a single record to correct; this
    # can, one entry per actual InvoiceItem.
    line_items = [{"id": it.id, "label": _invoice_master_bill_label(it.description), "description": it.description,
                   "amount": float(it.unit_price) * it.quantity, "reason": it.reason} for it in inv.items]

    return {
        "source": "invoice", "source_id": inv.id, "header": header, "mess": mess,
        "rows": assembled["rows"], "line_items": line_items, "last_debit_balance": float(inv.last_debit_balance or 0),
        "total_debit": assembled["total_debit"], "credits": assembled["credits"], "net_debits": assembled["net_debits"],
        "bill_serial_number": inv.bill_serial_number, "bill_made_at": inv.bill_made_at,
        "preset_heads": MASTER_BILL_PRESET_HEADS, "payment_breakdown": payment_breakdown,
        "status": inv.status.value,
    }


@router.post("/invoices/{invoice_id}/master-bill-lines")
async def add_master_bill_line(invoice_id: int, data: MasterBillLineCreate, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Adds a new line item to an already-generated invoice while it's still
    a draft in the Interactive Invoice Table (before "Make Bill" locks it) -
    the paper form's remaining preset heads (Masjid, Library, Sweeper...)
    that have no automated computation. Once bill_made_at is set, a change
    here would bypass Make Bill's lock, so it's blocked - a genuine
    correction after that point goes through the existing Manager-approved
    InvoiceEditRequest flow instead, unchanged."""
    if not check_permission(current_user, "billing", "create"):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.bill_made_at is not None:
        raise HTTPException(status_code=400, detail="This bill has been made - further changes require a correction request")

    item = InvoiceItem(invoice_id=inv.id, description=data.head.strip(), quantity=1, unit_price=data.amount, total_price=data.amount, reason=data.reason.strip())
    db.add(item)
    inv.subtotal = float(inv.subtotal) + data.amount
    inv.total_amount = float(inv.total_amount) + data.amount
    db.commit()
    db.refresh(item)
    log_audit(db, current_user.id, current_user.full_name, AuditAction.CREATE, "invoice_items", item.id,
              after_state=serialize_model(item), reason=item.reason, ip_address=request.client.host)
    return {"id": item.id, "description": item.description, "amount": float(item.unit_price)}


@router.put("/invoices/{invoice_id}/items/{item_id}/correct")
async def correct_master_bill_line(invoice_id: int, item_id: int, data: MasterBillLineCorrection, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Inline correction of an already-populated line in the pre-lock
    Interactive Master Bill Table (Workflow 1 step 3) - distinct from
    add_master_bill_line (adds a brand-new line) and from InvoiceEditRequest
    (the Manager-approval flow for a bill AFTER "Make Bill" locks it, still
    unchanged/blocked here). Any correction here auto-dispatches an Alert to
    the Manager's Alerts & Approvals inbox with the before/after and reason -
    no approval gate, since the bill isn't locked yet and the Clerk still
    owns it outright; this is a paper trail, not a checkpoint."""
    if not check_permission(current_user, "billing", "create"):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.bill_made_at is not None:
        raise HTTPException(status_code=400, detail="This bill has been made - further changes require a correction request")
    item = db.query(InvoiceItem).filter(InvoiceItem.id == item_id, InvoiceItem.invoice_id == invoice_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Line item not found on this invoice")

    old_amount = float(item.unit_price) * float(item.quantity)
    delta = data.new_amount - old_amount
    before = serialize_model(item)
    item.unit_price = data.new_amount / float(item.quantity or 1)
    item.total_price = data.new_amount
    inv.subtotal = float(inv.subtotal) + delta
    inv.total_amount = float(inv.total_amount) + delta
    db.commit()
    db.refresh(item)

    log_audit(db, current_user.id, current_user.full_name, AuditAction.UPDATE, "invoice_items", item.id,
              before_state=before, after_state=serialize_model(item), reason=data.correction_reason, ip_address=request.client.host)
    create_alert(
        db,
        title=f"Bill correction: {item.description}",
        message=(f"{current_user.full_name} corrected '{item.description}' on invoice {inv.invoice_number} "
                  f"from Rs {old_amount:,.2f} to Rs {data.new_amount:,.2f}. Reason: {data.correction_reason}"),
        severity=AlertSeverity.MEDIUM, module="billing", entity_type="invoice", entity_id=inv.id,
        detail={"item_id": item.id, "description": item.description, "original_amount": old_amount,
                "corrected_amount": data.new_amount, "reason": data.correction_reason, "corrected_by": current_user.full_name},
    )
    return {"id": item.id, "description": item.description, "amount": float(item.unit_price), "total_amount": float(inv.total_amount)}


@router.put("/invoices/{invoice_id}/last-debit-balance")
async def set_last_debit_balance(invoice_id: int, data: LastDebitBalanceUpdate, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "billing", "edit"):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.bill_made_at is not None:
        raise HTTPException(status_code=400, detail="This bill has been made - the carried-over balance can no longer be changed here")
    before = serialize_model(inv)
    inv.last_debit_balance = data.last_debit_balance
    db.commit()
    log_audit(db, current_user.id, current_user.full_name, AuditAction.UPDATE, "invoices", inv.id, before_state=before, after_state=serialize_model(inv), ip_address=request.client.host)
    return {"last_debit_balance": float(inv.last_debit_balance)}


@router.post("/invoices/{invoice_id}/make-bill")
async def make_bill(invoice_id: int, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Locks the Interactive Invoice Table (no more line additions or last
    debit balance edits) and enables Print. Idempotent - re-clicking on an
    already-made bill just returns its existing lock time."""
    if not check_permission(current_user, "billing", "edit"):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if inv.bill_made_at is None:
        inv.bill_made_at = datetime.utcnow()
        db.commit()
        log_audit(db, current_user.id, current_user.full_name, AuditAction.UPDATE, "invoices", inv.id, after_state={"bill_made_at": inv.bill_made_at.isoformat()}, ip_address=request.client.host)
    return {"bill_made_at": inv.bill_made_at}


@router.get("/invoices/{invoice_id}/payments")
async def list_payments(invoice_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    payments = db.query(InvoicePayment).filter(InvoicePayment.invoice_id == invoice_id).order_by(InvoicePayment.created_at.desc()).all()
    return [{"id": p.id, "invoice_id": p.invoice_id, "amount": float(p.amount), "method": p.method,
             "voucher_number": p.voucher_number, "notes": p.notes, "received_by": p.received_by, "created_at": p.created_at} for p in payments]


@router.post("/invoices/{invoice_id}/payments")
async def record_payment(invoice_id: int, data: PaymentCreate, request: Request, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if not check_permission(current_user, "clerk_desk", "create"):
        raise HTTPException(status_code=403, detail="Permission denied")
    if inv.status == InvoiceStatus.VOID:
        raise HTTPException(status_code=400, detail="Cannot record a payment against a void invoice")

    balance_due = _invoice_net_owed(inv) - float(inv.amount_paid or 0)
    if data.amount > balance_due + 0.01:  # small epsilon for float rounding
        raise HTTPException(status_code=400, detail=f"Payment of {data.amount:.2f} exceeds balance due of {balance_due:.2f}")

    before = serialize_model(inv)
    payment = InvoicePayment(
        invoice_id=invoice_id, amount=data.amount, method=data.method,
        voucher_number=data.voucher_number, notes=data.notes, received_by=current_user.id,
    )
    db.add(payment)

    inv.amount_paid = float(inv.amount_paid or 0) + data.amount
    _apply_invoice_payment_status(inv)

    db.commit()
    db.refresh(inv)
    db.refresh(payment)

    log_audit(db, current_user.id, current_user.full_name, AuditAction.UPDATE, "invoices", inv.id,
              before_state=before, after_state=serialize_model(inv),
              reason=f"Payment of {data.amount:.2f} recorded ({data.method or 'unspecified method'})", ip_address=request.client.host)
    return {"id": payment.id, "amount_paid": float(inv.amount_paid), "balance_due": _invoice_net_owed(inv) - float(inv.amount_paid), "status": inv.status.value}


@router.get("/payments/{payment_id}/receipt-data")
async def payment_receipt_data(payment_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Print data for the mess's cash-receipt format ('Received from ... the
    sum of Rupees ... on account of Mess Bill ... by Cash/Cheque'). Same gate
    as /desk - guest name and amount, and payment_id is directly enumerable."""
    if not (check_permission(current_user, "billing", "view") or check_permission(current_user, "clerk_desk", "view")):
        raise HTTPException(status_code=403, detail="Permission denied")
    payment = db.query(InvoicePayment).filter(InvoicePayment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    inv = payment.invoice
    booking = inv.booking if inv else None

    mess_name = get_setting_str(db, "mess_name", "EME Officers Mess")
    mess_address = get_setting_str(db, "mess_address", "204 Firdousi Road, Rawalpindi")
    mess_phone = get_setting_str(db, "mess_phone", "Tele No. G.H.Q 31725")

    guest = _guest_display_name(booking) if booking else (inv.guest.full_name if inv and inv.guest else "-")
    payload = "\n".join([
        mess_name,
        f"Receipt No: {payment.id}",
        f"Received from: {guest}",
        f"Amount: Rs {float(payment.amount):,.0f}",
        f"On account of: {inv.bill_type or 'mess'} bill {inv.invoice_number}" if inv else "",
        f"By: {payment.method or 'Cash'}",
        f"Date: {payment.created_at.strftime('%d-%m-%Y')}",
    ])

    return {
        "receipt_no": payment.id,
        "date": payment.created_at,
        "received_from": guest,
        "amount": float(payment.amount),
        "method": payment.method or "Cash",
        "notes": payment.notes,
        "on_account_of": f"{({'room': 'Room Bill', 'mess': 'Mess Bill'}.get(inv.bill_type or 'combined', 'Bill'))} {inv.invoice_number}" if inv else None,
        "invoice_number": inv.invoice_number if inv else None,
        "room_number": booking.room.room_number if booking and booking.room else None,
        "mess": {"name": mess_name, "address": mess_address, "phone": mess_phone},
        "qr_svg": _invoice_qr_svg(payload),
    }


@router.get("/dashboard-stats")
async def billing_stats(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Revenue totals for the Clerk dashboard - same gate as /desk."""
    if not (check_permission(current_user, "billing", "view") or check_permission(current_user, "clerk_desk", "view")):
        raise HTTPException(status_code=403, detail="Permission denied")
    today = date.today()
    month_start = today.replace(day=1)

    today_invoices = db.query(Invoice).filter(
        Invoice.issue_date == today,
        Invoice.status.in_(["issued", "paid"]),
    ).all()

    month_total = sum(float(inv.total_amount) for inv in db.query(Invoice).filter(
        Invoice.issue_date >= month_start,
        Invoice.status.in_(["issued", "paid"]),
    ).all())

    # Overdue is derived, not a stored status: an issued bill past its due
    # date with money still owing. (Nothing in the system ever transitions an
    # invoice to the stored OVERDUE status.)
    overdue = db.query(Invoice).filter(
        Invoice.status == InvoiceStatus.ISSUED, Invoice.due_date < today,
    ).count()

    # Cashier figures for the Clerk dashboard: what actually landed in hand
    # today (payments recorded today, split by method) - distinct from
    # today_revenue above, which is bills issued today (accrual, not cash).
    # InvoicePayment.created_at is stored via datetime.utcnow() (unlike
    # Invoice.issue_date, a local-date column set from date.today()) - the
    # "today" boundary here must use the same UTC clock it was written with,
    # or a payment made moments ago can fall outside it depending on the
    # server's UTC offset.
    utc_today = datetime.utcnow().date()
    utc_today_start = datetime.combine(utc_today, datetime.min.time())
    utc_month_start = datetime.combine(utc_today.replace(day=1), datetime.min.time())
    today_payments = db.query(InvoicePayment).filter(InvoicePayment.created_at >= utc_today_start).all()
    today_collections = sum(float(p.amount) for p in today_payments)
    # Month-to-date collections - same UTC boundary as today_collections since
    # payment timestamps are stored in UTC.
    month_collections = sum(float(p.amount) for p in db.query(InvoicePayment).filter(
        InvoicePayment.created_at >= utc_month_start).all())
    payment_methods_today: dict = {}
    for p in today_payments:
        method = p.method or "Cash"
        payment_methods_today[method] = payment_methods_today.get(method, 0.0) + float(p.amount)

    # Room vs mess mix for today's finalized bills - each invoice is one side
    # only since instant-checkout stopped merging both into one invoice;
    # legacy 'combined' rows from before that change aren't split further.
    today_room_revenue = sum(float(inv.total_amount) for inv in today_invoices if inv.bill_type == "room")
    today_mess_revenue = sum(float(inv.total_amount) for inv in today_invoices if inv.bill_type == "mess")

    # Discounts given today - a leakage-watch figure, since granting these is
    # the Clerk's own approval authority.
    today_discounts = sum(float(inv.discount) for inv in today_invoices if inv.discount)

    return {
        "today_revenue": sum(float(inv.total_amount) for inv in today_invoices),
        "today_invoice_count": len(today_invoices),
        "month_revenue": month_total,
        "overdue_invoices": overdue,
        "today_collections": round(today_collections, 2),
        "payment_methods_today": [
            {"method": m, "amount": round(a, 2)}
            for m, a in sorted(payment_methods_today.items(), key=lambda x: -x[1])
        ],
        "today_room_revenue": today_room_revenue,
        "today_mess_revenue": today_mess_revenue,
        "today_discounts": round(today_discounts, 2),
        "month_collections": round(month_collections, 2),
    }


def _period_bounds(period: str, date_str: Optional[str]):
    """Resolves a period=month|year + optional reference date into
    [start, end) date bounds, defaulting to the period containing today."""
    ref = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else date.today()
    if period == "year":
        start = ref.replace(month=1, day=1)
        end = start.replace(year=start.year + 1)
    else:
        start = ref.replace(day=1)
        end = date(start.year + 1, 1, 1) if start.month == 12 else start.replace(month=start.month + 1)
    return start, end


@router.get("/reports/summary")
async def billing_report_summary(period: str = Query("month", pattern="^(month|year)$"),
                                  as_of: Optional[str] = Query(None, alias="date", description="Any date within the period, defaults to today"),
                                  db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Clerk-facing month/year income-vs-cost report - reachable on the same
    billing:view permission the Clerk already has, so it doesn't need the
    Manager-only 'reports' module.

    Cost is only ever tracked on the mess/kitchen side (Procurement spend +
    WasteLog) - there is no housekeeping/utility/maintenance cost model for
    rooms anywhere in this system, so `room.cost`/`room.margin` come back
    `None` rather than a fabricated figure. Don't attribute any procurement
    or waste cost to the room side - every InventoryItem in this system is
    kitchen/mess stock, never room supplies."""
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    start, end = _period_bounds(period, as_of)
    start_dt = datetime.combine(start, datetime.min.time())
    end_dt = datetime.combine(end, datetime.min.time())

    invoices = db.query(Invoice).filter(Invoice.issue_date >= start, Invoice.issue_date < end).all()
    billed = [inv for inv in invoices if inv.status.value in ("issued", "paid")]
    voided = [inv for inv in invoices if inv.status.value == "void"]

    room_revenue = sum(float(inv.total_amount) for inv in billed if inv.bill_type == "room")
    mess_revenue = sum(float(inv.total_amount) for inv in billed if inv.bill_type == "mess")
    total_revenue = sum(float(inv.total_amount) for inv in billed)
    other_revenue = round(total_revenue - room_revenue - mess_revenue, 2)
    discounts_total = sum(float(inv.discount) for inv in invoices if inv.discount)

    today = date.today()
    overdue_amount = sum(
        max(float(inv.total_amount) - float(inv.amount_paid or 0), 0.0)
        for inv in invoices
        if inv.status.value == "issued" and inv.due_date and inv.due_date < today
    )

    # Cash in: money actually received during the period (payment date, not
    # invoice issue date - a payment can land against a bill issued earlier).
    payments = db.query(InvoicePayment).filter(
        InvoicePayment.created_at >= start_dt, InvoicePayment.created_at < end_dt,
    ).all()
    collections_total = sum(float(p.amount) for p in payments)
    collections_by_method: dict = {}
    for p in payments:
        m = p.method or "Cash"
        collections_by_method[m] = collections_by_method.get(m, 0.0) + float(p.amount)

    # Mess cost: self-purchase stock spend + logged waste in the period -
    # the same two figures Manager's dashboard treats as "cost", both of
    # which are entirely kitchen/mess stock. Not an actual paid-to-vendor
    # cash ledger (no such thing exists), but a real spend figure.
    procurement_spend = float(db.query(func.sum(StockBatch.quantity * StockBatch.unit_cost)).filter(
        StockBatch.created_at >= start_dt, StockBatch.created_at < end_dt,
    ).scalar() or 0)
    waste_cost = float(db.query(func.sum(WasteLog.cost)).filter(
        WasteLog.created_at >= start_dt, WasteLog.created_at < end_dt,
    ).scalar() or 0)
    mess_cost = procurement_spend + waste_cost

    return {
        "period": period,
        "start_date": start.isoformat(),
        "end_date": (end - timedelta(days=1)).isoformat(),
        "room": {
            "income": round(room_revenue, 2),
            "cost": None,
            "margin": None,
        },
        "mess": {
            "income": round(mess_revenue, 2),
            "cost": round(mess_cost, 2),
            "margin": round(mess_revenue - mess_cost, 2),
        },
        "mess_cost_breakdown": {
            "procurement": round(procurement_spend, 2),
            "waste": round(waste_cost, 2),
        },
        "other_revenue": max(other_revenue, 0.0),
        "total_revenue": round(total_revenue, 2),
        "invoice_count": len(billed),
        "discounts_total": round(discounts_total, 2),
        "void_count": len(voided),
        "void_amount": round(sum(float(inv.total_amount) for inv in voided), 2),
        "overdue_amount": round(overdue_amount, 2),
        "cash_in_total": round(collections_total, 2),
        "cash_in_by_method": [
            {"method": m, "amount": round(a, 2)} for m, a in sorted(collections_by_method.items(), key=lambda x: -x[1])
        ],
    }


@router.get("/stock-summary")
async def billing_stock_summary(limit: int = Query(8, ge=1, le=50),
                                 db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Read-only stock valuation snapshot for the Clerk's billing reports.
    Deliberately narrow (no inventory:view grant) - Inventory management
    itself stays Kitchen NCO's job; this only surfaces the total value and a
    top-items list a Clerk needs for the monthly report."""
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    items = db.query(InventoryItem).filter(InventoryItem.is_active == True).all()
    rows = []
    total_value = 0.0
    low_stock_count = 0
    for item in items:
        batches = db.query(StockBatch).filter(StockBatch.item_id == item.id, StockBatch.is_active == True).all()
        quantity = sum(b.quantity for b in batches)
        value = sum(float(b.unit_cost or 0) * b.quantity for b in batches)
        total_value += value
        if item.reorder_level and quantity <= item.reorder_level:
            low_stock_count += 1
        rows.append({"id": item.id, "name": item.name, "sku": item.sku, "unit": item.unit,
                      "quantity": quantity, "value": round(value, 2)})
    rows.sort(key=lambda r: r["value"], reverse=True)
    return {
        "total_stock_value": round(total_value, 2),
        "low_stock_count": low_stock_count,
        "top_items": rows[:limit],
    }


@router.get("/export/invoices")
async def export_invoices(start: Optional[str] = Query(None), end: Optional[str] = Query(None),
                           db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Excel export of invoices, optionally bounded by issue_date. Gated on
    billing:view (not the import_export module) so the Clerk can export their
    own domain's data without the master-data import/export permission."""
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    q = db.query(Invoice).options(joinedload(Invoice.booking).joinedload(Booking.room), joinedload(Invoice.guest))
    if start:
        q = q.filter(Invoice.issue_date >= datetime.strptime(start, "%Y-%m-%d").date())
    if end:
        q = q.filter(Invoice.issue_date <= datetime.strptime(end, "%Y-%m-%d").date())
    invoices = q.order_by(Invoice.issue_date.desc(), Invoice.id.desc()).limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append(["Invoice #", "Type", "Guest", "Room", "Issue Date", "Due Date", "Status", "Total", "Paid", "Balance", "Discount"])
    for inv in invoices:
        b = inv.booking
        guest_name = b.guest_name if b else (inv.guest.full_name if inv.guest else "")
        room_no = b.room.room_number if b and b.room else ""
        ws.append([
            inv.invoice_number, inv.bill_type or "combined", guest_name, room_no,
            inv.issue_date.isoformat() if inv.issue_date else "", inv.due_date.isoformat() if inv.due_date else "",
            inv.status.value, float(inv.total_amount), float(inv.amount_paid or 0),
            float(inv.total_amount) - float(inv.amount_paid or 0), float(inv.discount or 0),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename=invoices_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"})


@router.get("/export/payments")
async def export_payments(start: Optional[str] = Query(None), end: Optional[str] = Query(None),
                           db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Excel export of recorded payments (cash-in ledger), optionally bounded
    by the payment's received date."""
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    q = db.query(InvoicePayment).options(
        joinedload(InvoicePayment.invoice).joinedload(Invoice.booking).joinedload(Booking.room),
        joinedload(InvoicePayment.invoice).joinedload(Invoice.guest),
    )
    if start:
        q = q.filter(InvoicePayment.created_at >= datetime.strptime(start, "%Y-%m-%d"))
    if end:
        q = q.filter(InvoicePayment.created_at < datetime.strptime(end, "%Y-%m-%d") + timedelta(days=1))
    payments = q.order_by(InvoicePayment.created_at.desc()).limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws.append(["Date", "Invoice #", "Guest", "Room", "Amount", "Method", "Notes"])
    for p in payments:
        inv = p.invoice
        b = inv.booking if inv else None
        guest_name = b.guest_name if b else (inv.guest.full_name if inv and inv.guest else "")
        room_no = b.room.room_number if b and b.room else ""
        ws.append([
            p.created_at.isoformat() if p.created_at else "", inv.invoice_number if inv else "",
            guest_name, room_no, float(p.amount), p.method or "", p.notes or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename=payments_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"})


def _ag_branch_query(db: Session, voucher_number: str, date_from: Optional[str], date_to: Optional[str]):
    q = db.query(InvoicePayment).options(
        joinedload(InvoicePayment.invoice).joinedload(Invoice.booking),
    ).filter(InvoicePayment.ag_branch_fee.isnot(None))
    if voucher_number:
        q = q.filter(InvoicePayment.voucher_number.contains(voucher_number))
    if date_from:
        q = q.filter(InvoicePayment.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        q = q.filter(InvoicePayment.created_at < datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
    return q.order_by(InvoicePayment.created_at.desc())


@router.get("/ag-branch-advances")
async def list_ag_branch_advances(
    voucher_number: str = "", date_from: str = "", date_to: str = "",
    page: int = Query(1, ge=1), page_size: int = Query(25, ge=1, le=100),
    db: Session = Depends(get_db), current_user=Depends(get_current_user),
):
    """Online-booking advance payments with their AG Branch 10% deduction
    broken out - voucher search + the four figures the spec requires
    (gross/fee/net/customer). Every row here is strictly an online-advance
    payment (ag_branch_fee is only ever set on those, see instant_checkout)."""
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    q = _ag_branch_query(db, voucher_number, date_from, date_to)
    total = q.count()
    payments = q.offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for p in payments:
        inv = p.invoice
        b = inv.booking if inv else None
        gross = float(p.amount)
        fee = float(p.ag_branch_fee or 0)
        items.append({
            "payment_id": p.id, "voucher_number": p.voucher_number,
            "customer_name": _guest_display_name(b) if b else None,
            "booking_reference": b.booking_reference if b else None,
            "invoice_number": inv.invoice_number if inv else None,
            "gross_advance": gross, "ag_branch_fee": fee, "net_amount": gross - fee,
            "date": p.created_at,
        })
    return {"items": items, "total": total}


@router.get("/export/ag-branch-advances")
async def export_ag_branch_advances(
    voucher_number: str = "", date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), current_user=Depends(get_current_user),
):
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    payments = _ag_branch_query(db, voucher_number, date_from, date_to).limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "AG Branch Advances"
    ws.append(["Voucher Number", "Customer Name & Rank", "Gross Advance Received", "10% AG Branch Fee Deducted", "Net Amount Received by Mess", "Date", "Booking Ref"])
    for p in payments:
        inv = p.invoice
        b = inv.booking if inv else None
        gross = float(p.amount)
        fee = float(p.ag_branch_fee or 0)
        ws.append([
            p.voucher_number or "", _guest_display_name(b) if b else "",
            gross, fee, gross - fee,
            p.created_at.isoformat() if p.created_at else "", b.booking_reference if b else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename=ag_branch_advances_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"})


@router.get("/export/ag-branch-advances/pdf")
async def export_ag_branch_advances_pdf(
    voucher_number: str = "", date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), current_user=Depends(get_current_user),
):
    if not check_permission(current_user, "billing", "view"):
        raise HTTPException(status_code=403, detail="Permission denied")
    payments = _ag_branch_query(db, voucher_number, date_from, date_to).limit(5000).all()

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    rows = [["Voucher Number", "Customer Name & Rank", "Gross Advance\nReceived", "10% AG Branch\nFee Deducted", "Net Amount\nReceived by Mess", "Date", "Booking Ref"]]
    gross_total = fee_total = net_total = 0.0
    for p in payments:
        inv = p.invoice
        b = inv.booking if inv else None
        gross = float(p.amount)
        fee = float(p.ag_branch_fee or 0)
        net = gross - fee
        gross_total += gross
        fee_total += fee
        net_total += net
        rows.append([
            p.voucher_number or "-", _guest_display_name(b) if b else "-",
            f"{gross:,.2f}", f"{fee:,.2f}", f"{net:,.2f}",
            p.created_at.strftime("%d-%m-%Y") if p.created_at else "-", b.booking_reference if b else "-",
        ])
    rows.append(["", "Total", f"{gross_total:,.2f}", f"{fee_total:,.2f}", f"{net_total:,.2f}", "", ""])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(get_setting_str(db, "mess_name", "EME Officers Mess"), styles["Title"]),
        Paragraph("AG Branch Advance Report", styles["Heading2"]),
        Spacer(1, 8),
    ]
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (2, 0), (4, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f3f4f6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": f"attachment; filename=ag_branch_advances_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"})
